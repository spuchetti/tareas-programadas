import json
import os
import time
import base64
from datetime import datetime
from zoneinfo import ZoneInfo
from io import BytesIO
from copy import copy
import concurrent.futures
import traceback
import zipfile
import tempfile
import shutil
import xml.etree.ElementTree as ET

import openpyxl

# Drive API
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from googleapiclient.errors import HttpError

# Gmail API
from google.oauth2.credentials import Credentials as GmailCreds
from googleapiclient.discovery import build as gmail_build
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

# ======================================
# CONFIGURACIÓN
# ======================================

TEXTO_OBJETIVO = "REVISAR"
COLUMNA_OBJETIVO = 33
FILA_INICIO = 4
FOLDER_ID = "1_Xb2jrtr3Sjwi8-2nhT2k53KZ6CLE5hJ"

MAXIMO_HILOS = 5
INTENTOS_MAX = 3
ESPERA_REINTENTO = 5

# ======================================
# UTILIDADES
# ======================================

def nombre_mes(numero):
    meses = {
        "01": "Enero", "02": "Febrero", "03": "Marzo", "04": "Abril",
        "05": "Mayo", "06": "Junio", "07": "Julio", "08": "Agosto",
        "09": "Septiembre", "10": "Octubre", "11": "Noviembre", "12": "Diciembre"
    }
    return meses.get(str(numero), "???")

# ======================================
# SANITIZAR XML / EXCEL
# ======================================

def sanitizar_libro_remover_filtros(fh):
    """
    Intenta reparar un archivo .xlsx o .xlsm eliminando nodos <autoFilter>
    y otros nodos problemáticos. Devuelve BytesIO con el archivo reempaquetado
    o None si no pudo repararlo.
    """
    try:
        tmpdir = tempfile.mkdtemp()
        ruta_entrada = os.path.join(tmpdir, "input_file")
        with open(ruta_entrada, "wb") as f:
            f.write(fh.getvalue())

        ruta_extraida = os.path.join(tmpdir, "extracted")
        os.makedirs(ruta_extraida, exist_ok=True)

        with zipfile.ZipFile(ruta_entrada, "r") as zin:
            zin.extractall(ruta_extraida)

        ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        worksheets_dir = os.path.join(ruta_extraida, "xl", "worksheets")

        if os.path.isdir(worksheets_dir):
            for nombre_arch in os.listdir(worksheets_dir):
                if not nombre_arch.lower().endswith(".xml"):
                    continue
                ruta = os.path.join(worksheets_dir, nombre_arch)
                try:
                    tree = ET.parse(ruta)
                    root = tree.getroot()

                    modificado = False
                    for auto in root.findall("main:autoFilter", ns):
                        root.remove(auto)
                        modificado = True

                    for ext in root.findall("main:extLst", ns):
                        root.remove(ext)
                        modificado = True

                    for fcol in root.findall(".//main:filterColumn", ns):
                        for filt in fcol.findall("main:filters", ns):
                            fcol.remove(filt)
                            modificado = True

                    if modificado:
                        tree.write(ruta, encoding="utf-8", xml_declaration=True)
                except ET.ParseError:
                    # Si no se puede parsear una hoja, seguimos con las demás
                    pass

        ruta_salida = os.path.join(tmpdir, "output_file.xlsx")
        with zipfile.ZipFile(ruta_salida, "w", zipfile.ZIP_DEFLATED) as zout:
            for root_dir, dirs, files in os.walk(ruta_extraida):
                for file in files:
                    fullpath = os.path.join(root_dir, file)
                    arcname = os.path.relpath(fullpath, ruta_extraida)
                    zout.write(fullpath, arcname)

        with open(ruta_salida, "rb") as f:
            datos = f.read()

        shutil.rmtree(tmpdir)
        return BytesIO(datos)

    except Exception as e:
        try:
            shutil.rmtree(tmpdir)
        except Exception:
            pass
        print(f"❌ sanitizar_libro_remover_filtros error: {e}")
        return None

# ======================================
# REGISTROS (LOGS)
# ======================================

def registrar_inicio():
    ahora = datetime.now(ZoneInfo("America/Argentina/Buenos_Aires"))
    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print("🚀 INICIO DEL PROCESO AUTOMÁTICO")
    print(f"📅 Fecha y hora: {ahora.strftime('%Y-%m-%d %H:%M:%S')}")
    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")


def registrar_resumen(archivos, encontrados, inicio):
    duracion = time.time() - inicio
    print("\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print("📊 RESUMEN FINAL")
    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print(f"🗂 Archivos procesados: {len(archivos)}")
    print(f"🔎 Archivos detectados: {len(encontrados)}")
    print(f"⏱ Tiempo total: {duracion:.2f} segundos")
    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")

    if encontrados:
        print("📌 Archivos encontrados:")
        for f in encontrados:
            print(f"   ✔ {f}")
    else:
        print("✅ No se encontraron archivos con diferencias")

    print("\n✔ Proceso completado.\n")

# ======================================
# HTML PARA EMAIL
# ======================================

def generar_html(periodo, procesados, detectados, lista, fecha):
    if lista:
        lista_html = "\n".join(f"<li>✔ {os.path.splitext(item)[0]}</li>" for item in lista)
    else:
        lista_html = "<li>No se encontraron archivos</li>"

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width,initial-scale=1"/>
  <title>Resultado</title>
</head>

<body style="font-family: Arial, Helvetica, sans-serif; color:#222; line-height:1.4; padding:18px;">

  <div style="
      background: linear-gradient(90deg,#0a7bdc,#16a085);
      padding: 18px;
      border-radius: 8px;
      color: white;
      margin-bottom: 18px;
    ">
    <h2 style="margin:0;">🟢🔵 OSER - CONTROL AUTOMÁTICO FONDO VOLUNTARIO</h2>
    <div style="opacity:0.9; font-size:14px; margin-top:4px;">Resultado de la revisión automática</div>
  </div>

  <p><strong>Periodo:</strong> {periodo}</p>
  <p>
    <strong>Archivos procesados:</strong> {procesados}<br/>
    <strong>Total detectados:</strong> {detectados}
  </p>

  <hr style="margin:18px 0;">

  <p><strong>Reparticiones:</strong></p>
  <ul>
    {lista_html}
  </ul>

  <hr style="margin:18px 0;">

  <p style="font-size:0.9em; color:#555;">
    Generado: {fecha}
  </p>

<div style="text-align:right; margin-top:25px;">
    <img src="https://raw.githubusercontent.com/spuchetti/tareas-programadas/main/assets/robot.jpg"
         width="140"
         style="opacity:0.55; display:inline-block;"/>
  </div>

</body>
</html>"""

    return html

# ======================================
# EMAIL - enviar con adjuntos
# ======================================

def enviar_email_html_con_adjuntos(asunto, html, lista_adjuntos):
    token_json = os.getenv("GMAIL_TOKEN")
    mail_to = os.getenv("SMTP_TO")

    if not token_json or not mail_to:
        print("⚠ NO se enviará email: faltan credenciales.")
        return

    creds = GmailCreds.from_authorized_user_info(json.loads(token_json))

    try:
        servicio = gmail_build("gmail", "v1", credentials=creds)

        msg = MIMEMultipart()
        msg["To"] = mail_to
        msg["From"] = "me"
        msg["Subject"] = asunto

        msg.attach(MIMEText(html, "html", "utf-8"))

        for ruta in lista_adjuntos:
            with open(ruta, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", f'attachment; filename="{os.path.basename(ruta)}"')
                msg.attach(part)

        encoded = base64.urlsafe_b64encode(msg.as_bytes()).decode("utf-8")
        servicio.users().messages().send(userId="me", body={"raw": encoded}).execute()

        print("📧 Email enviado correctamente con adjuntos.")

    except Exception as e:
        print(f"❌ Error enviando email: {e}")
        traceback.print_exc()

# ======================================
# GOOGLE DRIVE
# ======================================

def inicializar_drive():
    try:
        cfg = json.loads(os.getenv("GDRIVE_JSON"))  # NO RENOMBRAR VARIABLE DE ENTORNO
        creds = Credentials.from_service_account_info(cfg)
        servicio = build("drive", "v3", credentials=creds, cache_discovery=False)
        return servicio
    except Exception as e:
        print(f"❌ Error iniciando Drive: {e}")
        traceback.print_exc()
        return None


def request_drive_con_reintentos(funcion, descripcion):
    for intento in range(INTENTOS_MAX):
        try:
            return funcion()
        except HttpError as e:
            if e.resp.status in [403, 500, 503]:
                print(f"⏳ Error {descripcion}, reintento {intento+1}/{INTENTOS_MAX}")
                time.sleep(ESPERA_REINTENTO)
                continue
            return None
        except Exception:
            return None
    return None


def obtener_archivos(desde_drive):
    print("📁 Buscando archivos en Drive...\n")
    consulta = f"'{FOLDER_ID}' in parents and trashed=false"

    res = request_drive_con_reintentos(
        lambda: desde_drive.files().list(
            q=consulta,
            pageSize=200,
            fields="files(id,name,mimeType)",
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        ).execute(),
        "listar archivos"
    )

    if not res:
        return []

    archivos = res.get("files", [])

    archivos_validos = []
    for a in archivos:
        nombre = a["name"].lower()

        if (
            nombre.endswith(".xlsx") or
            nombre.endswith(".xlsm") or
            a["mimeType"] == "application/vnd.google-apps.spreadsheet"
        ):
            archivos_validos.append(a)

    return archivos_validos


def descargar_archivo(desde_drive, archivo):
    file_id = archivo["id"]
    mime = archivo["mimeType"]

    try:
        if mime == "application/vnd.google-apps.spreadsheet":
            req = desde_drive.files().export_media(
                fileId=file_id,
                mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            req = desde_drive.files().get_media(fileId=file_id)

        fh = BytesIO()
        downloader = MediaIoBaseDownload(fh, req)
        terminado = False

        while not terminado:
            _, terminado = downloader.next_chunk()

        fh.seek(0)
        return fh

    except Exception as e:
        print(f"❌ Error descargando {archivo['name']}: {e}")
        traceback.print_exc()
        return None

# ======================================
# PROCESAR EXCEL
# ======================================

def buscar_en_hoja(fh, nombre, hoja):
    try:
        fh.seek(0)
        wb = openpyxl.load_workbook(fh, data_only=True, read_only=True)

        if hoja not in wb.sheetnames:
            wb.close()
            return False

        ws = wb[hoja]

        for row in ws.iter_rows(min_row=FILA_INICIO, values_only=True):
            if row is None:
                break

            if len(row) < COLUMNA_OBJETIVO:
                continue

            val = row[COLUMNA_OBJETIVO - 1]

            if val is None:
                break

            if TEXTO_OBJETIVO.upper() in str(val).upper():
                wb.close()
                return True

        wb.close()
        return False

    except Exception as e:
        print(f"❌ Error leyendo {nombre}: {e}")
        traceback.print_exc()
        return False

# ======================================
# GENERAR ARCHIVO CON SOLO HOJAS NECESARIAS
# ======================================

def generar_archivo_solo_hojas(fh, nombre_archivo, hoja_periodo):
    try:
        fh.seek(0)

        libro_original = None
        intento_sanitizar = False

        for intento in range(2):
            try:
                libro_original = openpyxl.load_workbook(
                    fh,
                    data_only=False,
                    keep_vba=True,
                    keep_links=False
                )
                break
            except Exception as e:
                if not intento_sanitizar:
                    print(f"⚠ load_workbook falló: {e}. Intentando sanitizar XML (quitar filtros)...")
                    fh_sanitizado = sanitizar_libro_remover_filtros(fh)
                    intento_sanitizar = True
                    if fh_sanitizado is None:
                        print(f"⚠ No se pudo sanitizar: {nombre_archivo}")
                        break
                    else:
                        fh = fh_sanitizado
                        continue
                else:
                    print(f"⚠ Después de sanitizar, load_workbook sigue fallando: {e}")
                    libro_original = None
                    break

        if libro_original is None:
            try:
                fh.seek(0)
                wb_ro = openpyxl.load_workbook(fh, data_only=True, read_only=True)
                print(f"⚠ Usando fallback read_only para: {nombre_archivo} (pierde estilos/formulas)")
                wb_ro.close()
                return None
            except Exception:
                print(f"❌ No se pudo abrir siquiera en modo read_only: {nombre_archivo}")
                return None

        hojas_a_copiar = [hoja_periodo]

        if "MINIMOS" in libro_original.sheetnames:
            hojas_a_copiar.append("MINIMOS")

        nuevo_wb = openpyxl.Workbook()
        if nuevo_wb.active and nuevo_wb.active.title == "Sheet":
            nuevo_wb.remove(nuevo_wb.active)

        for hoja_nombre in hojas_a_copiar:
            origen = libro_original[hoja_nombre]
            nueva = nuevo_wb.create_sheet(title=hoja_nombre)

            for row in origen.iter_rows():
                for cell in row:
                    nueva_cell = nueva.cell(row=cell.row, column=cell.column)

                    if cell.data_type == 'f' and cell.value:
                        nueva_cell.value = f"={cell.value}"
                    else:
                        nueva_cell.value = cell.value

                    try:
                        nueva_cell.font = copy(cell.font)
                        nueva_cell.fill = copy(cell.fill)
                        nueva_cell.border = copy(cell.border)
                        nueva_cell.alignment = copy(cell.alignment)
                        nueva_cell.number_format = copy(cell.number_format)
                        nueva_cell.protection = copy(cell.protection)
                    except Exception:
                        pass

                    if cell.hyperlink:
                        nueva_cell.hyperlink = cell.hyperlink
                    if cell.comment:
                        try:
                            nueva_cell.comment = copy(cell.comment)
                        except Exception:
                            pass

            for col_letter, dim in origen.column_dimensions.items():
                try:
                    nueva.column_dimensions[col_letter].width = dim.width
                    nueva.column_dimensions[col_letter].hidden = dim.hidden
                except Exception:
                    pass
            for row_num, dim in origen.row_dimensions.items():
                try:
                    nueva.row_dimensions[row_num].height = dim.height
                    nueva.row_dimensions[row_num].hidden = dim.hidden
                except Exception:
                    pass

            try:
                nueva.page_setup = copy(origen.page_setup)
            except Exception:
                pass
            try:
                nueva.print_title_cols = copy(origen.print_title_cols)
                nueva.print_title_rows = copy(origen.print_title_rows)
            except Exception:
                pass

            nueva.auto_filter = None

        nombre_final = os.path.splitext(nombre_archivo)[0] + ".xlsx"
        carpeta = os.path.join("generados", os.path.splitext(nombre_archivo)[0])
        os.makedirs(carpeta, exist_ok=True)
        ruta = os.path.join(carpeta, nombre_final)

        nuevo_wb.save(ruta)
        nuevo_wb.close()
        libro_original.close()

        print(f"   → Archivo generado: {ruta}")
        return ruta

    except Exception as e:
        print(f"❌ Error generando archivo {nombre_archivo}: {e}")
        traceback.print_exc()
        return None

# ======================================
# PROCESAR UN ARCHIVO INDIVIDUAL
# ======================================

def procesar_archivo(archivo, hoja_periodo):
    try:
        servicio_drive = inicializar_drive()
        if not servicio_drive:
            return archivo["name"], None, None

        fh = descargar_archivo(servicio_drive, archivo)
        if not fh:
            return archivo["name"], None, None

        encontrado = buscar_en_hoja(fh, archivo["name"], hoja_periodo)
        fh_copia = BytesIO(fh.getvalue())

        return archivo["name"], encontrado, fh_copia
    except Exception as e:
        print(f"❌ Error procesando {archivo['name']}: {e}")
        traceback.print_exc()
        return archivo["name"], None, None

# ======================================
# PRINCIPAL
# ======================================

def ejecutar_principal():
    inicio = time.time()
    registrar_inicio()

    drive = inicializar_drive()
    if not drive:
        return

    archivos = obtener_archivos(drive)
    if not archivos:
        print("❌ No se encontraron archivos.")
        return

    ahora = datetime.now(ZoneInfo("America/Argentina/Buenos_Aires"))
    mes_num = ahora.month - 1 or 12
    periodo = f"{mes_num:02d}"
    periodo_legible = f"{nombre_mes(periodo)}/{ahora.year}"
    periodo_legible_upper = periodo_legible.upper()

    print(f"📄 Hoja a controlar: {periodo} ({periodo_legible})\n")

    encontrados = []
    adjuntos = []

    with concurrent.futures.ThreadPoolExecutor(max_workers=MAXIMO_HILOS) as pool:
        tareas = [
            pool.submit(procesar_archivo, a, periodo)
            for a in archivos
        ]

        for future in concurrent.futures.as_completed(tareas):
            nombre, ok, fh_copy = future.result()

            if ok:
                print(f"   ✔ {nombre} → REVISAR")
                encontrados.append(nombre)

                if fh_copy:
                    generado = generar_archivo_solo_hojas(fh_copy, nombre, periodo)
                    if generado:
                        adjuntos.append(generado)
            else:
                print(f"   ✖ {nombre}")

    html = generar_html(
        periodo_legible,
        len(archivos),
        len(encontrados),
        encontrados,
        ahora.strftime("%d-%m-%Y %H:%M:%S")
    )

    asunto = f"🟢🔵 OSER - CONTROL AUTOMÁTICO FONDO VOLUNTARIO | PERIODO: {periodo_legible_upper}"
    enviar_email_html_con_adjuntos(asunto, html, adjuntos)

    registrar_resumen(archivos, encontrados, inicio)


if __name__ == "__main__":
    ejecutar_principal()
