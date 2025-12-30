"""
Funciones para trabajar con archivos Excel
"""

import openpyxl
import zipfile
import tempfile
import shutil
import xml.etree.ElementTree as ET
from datetime import datetime
from zoneinfo import ZoneInfo
from io import BytesIO
import os


def obtener_hoja_mes_anterior():
    """Obtiene el nombre de la hoja del mes anterior (MM)"""
    ahora = datetime.now(ZoneInfo("America/Argentina/Buenos_Aires"))
    mes_anterior = ahora.month - 1 or 12
    return f"{mes_anterior:02d}"


def eliminar_tildes(texto):
    """
    Elimina tildes y caracteres especiales de un texto.
    Convierte caracteres acentuados a su equivalente sin tilde.
    
    Ejemplos:
    - "María González" → "Maria Gonzalez"
    - "José Pérez" → "Jose Perez"
    - "Ñandú" → "Nandu"
    - "Über" → "Uber"
    """
    if texto is None:
        return ""
    
    # Convertir a string si no lo es
    if not isinstance(texto, str):
        texto = str(texto)
    
    # Limpiar espacios
    texto = texto.strip()
    
    # Si está vacío después de limpiar
    if not texto:
        return ""
    
    # Diccionario de reemplazos para tildes y caracteres especiales
    reemplazos = {
        # Vocales minúsculas con tilde
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
        'à': 'a', 'è': 'e', 'ì': 'i', 'ò': 'o', 'ù': 'u',
        'ä': 'a', 'ë': 'e', 'ï': 'i', 'ö': 'o', 'ü': 'u',
        'â': 'a', 'ê': 'e', 'î': 'i', 'ô': 'o', 'û': 'u',
        
        # Vocales mayúsculas con tilde
        'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U',
        'À': 'A', 'È': 'E', 'Ì': 'I', 'Ò': 'O', 'Ù': 'U',
        'Ä': 'A', 'Ë': 'E', 'Ï': 'I', 'Ö': 'O', 'Ü': 'U',
        'Â': 'A', 'Ê': 'E', 'Î': 'I', 'Ô': 'O', 'Û': 'U',
        
        # Otros caracteres especiales
        'º': '°'
    }
    
    # Aplicar reemplazos
    resultado = ""
    for caracter in texto:
        if caracter in reemplazos:
            resultado += reemplazos[caracter]
        else:
            resultado += caracter
    
    # Asegurar que el texto sea compatible con CSV (| como separador)
    # Reemplazar el separador de CSV si aparece en el texto
    resultado = resultado.replace('|', '-')
    
    return resultado


def normalizar_texto(texto, eliminar_tildes_param=True):
    """
    Normaliza texto para CSV.
    
    Args:
        texto: Texto a normalizar
        eliminar_tildes_param: Si True, elimina tildes. Si False, las mantiene.
                               Por defecto es True para columnas D y H.
    """
    if texto is None:
        return ""
    
    # Convertir a string si no lo es
    if not isinstance(texto, str):
        texto = str(texto)
    
    # Limpiar espacios
    texto = texto.strip()
    
    # Si está vacío después de limpiar
    if not texto:
        return ""
    
    # SOLO si se solicita eliminar tildes
    if eliminar_tildes_param:
        return eliminar_tildes(texto)
    else:
        # Mantener tildes, solo reemplazar separador
        return texto.replace('|', '-')


def extraer_datos_excel(fh, nombre_archivo):
    """
    Extrae datos de un archivo Excel desde fila 4, columnas A a X
    en la hoja del mes anterior.
    Se detiene cuando encuentra "-" o celda vacía en columna A.
    
    NOTA: Para archivos de "Caja" empieza desde fila 5.
    
    COLUMNAS ESPECIALES:
    - D (4): Nombre y apellido - ELIMINAR TILDES (SÍ eliminar)
    - H (8): Repartición - ELIMINAR TILDES (SÍ eliminar)
    - I-X (9-24): Números con 2 decimales
    - Otras columnas de texto: mantener tildes
    """
    try:
        # Obtener hoja del mes anterior
        hoja_mes = obtener_hoja_mes_anterior()
        
        # Cargar libro
        fh.seek(0)
        wb = openpyxl.load_workbook(fh, data_only=True, read_only=True)
        
        # Verificar si existe la hoja del mes
        if hoja_mes not in wb.sheetnames:
            print(f"⚠ Hoja '{hoja_mes}' no encontrada en {nombre_archivo}")
            wb.close()
            return []
        
        ws = wb[hoja_mes]
        datos_extraidos = []
        
        # DETERMINAR FILA DE INICIO
        # Si es archivo de "Caja", empezar desde fila 5, sino desde fila 4
        if "caja" in nombre_archivo.lower():
            fila_inicio = 5
            print(f"   ⚙ Archivo 'Caja' detectado. Iniciando desde fila {fila_inicio}")
        else:
            fila_inicio = 4
        
        # Extraer datos desde fila determinada hasta el final, columnas A-X (1-24)
        for row_idx, row in enumerate(ws.iter_rows(min_row=fila_inicio, max_col=24, values_only=True), start=fila_inicio):
            # VERIFICAR CONDICIÓN DE PARADA: "-" o celda vacía en columna A
            primera_celda = row[0] if len(row) > 0 else None
            
            # Convertir a string para la verificación
            if primera_celda is None:
                primera_celda_str = ""
            elif isinstance(primera_celda, datetime):
                primera_celda_str = primera_celda.strftime("%Y-%m-%d")
            else:
                primera_celda_str = str(primera_celda).strip()
            
            # Detener si la primera celda es "-" o está vacía
            if primera_celda_str == "-" or primera_celda_str == "":
                print(f"   ⏹ Marcador de fin encontrado en fila {row_idx} ('{primera_celda_str}'). Fin de extracción.")
                break
            
            # También verificar si todas las celdas de la fila están vacías
            if all(cell is None or cell == '' or str(cell).strip() == '' for cell in row):
                print(f"   ⏹ Fila {row_idx} completamente vacía. Fin de extracción.")
                break
            
            # LIMPIAR Y FORMATAR CADA CELDA
            fila_limpia = []
            for col_idx, cell in enumerate(row, start=1):
                if cell is None:
                    fila_limpia.append("")
                elif isinstance(cell, datetime):
                    fila_limpia.append(cell.strftime("%Y-%m-%d"))
                else:
                    # Columnas A-H (1-8): formatear como texto
                    if col_idx <= 8:  
                        # Para columnas de texto
                        if isinstance(cell, (int, float)):
                            # Si es número (CUIL/DNI), eliminar decimales innecesarios
                            if isinstance(cell, float) and cell.is_integer():
                                fila_limpia.append(str(int(cell)))
                            else:
                                cell_str = str(cell)
                                # Eliminar .0 final si existe
                                if cell_str.endswith('.0'):
                                    cell_str = cell_str[:-2]
                                fila_limpia.append(cell_str)
                        else:
                            # Para texto
                            cell_str = str(cell) if cell is not None else ""
                            
                            # IMPORTANTE: Columnas D (4) y H (8) - ELIMINAR TILDES
                            if col_idx == 4 or col_idx == 8:
                                # COLUMNA D (Nombre) y H (Repartición) - ELIMINAR TILDES
                                fila_limpia.append(normalizar_texto(cell_str, eliminar_tildes_param=True))
                            else:
                                # Otras columnas de texto - mantener tildes
                                fila_limpia.append(normalizar_texto(cell_str, eliminar_tildes_param=False))
                    
                    else:  # Columnas I-X (9-24) - números SIEMPRE con 2 decimales y punto
                        if cell == "" or cell is None:
                            # Para celdas vacías, poner "0.00"
                            fila_limpia.append("0.00")
                        elif isinstance(cell, (int, float)):
                            # Para números enteros o decimales, formatear siempre con 2 decimales
                            formatted = f"{float(cell):.2f}"
                            fila_limpia.append(formatted)
                        else:
                            # Si es texto, intentar convertir a número
                            cell_str = str(cell).strip()
                            
                            if cell_str == "" or cell_str.lower() == "nan":
                                fila_limpia.append("0.00")
                            elif cell_str == "0":
                                fila_limpia.append("0.00")
                            else:
                                try:
                                    # Normalizar: reemplazar comas por puntos
                                    cell_normalized = cell_str.replace(',', '.')
                                    
                                    # Manejar múltiples puntos
                                    if cell_normalized.count('.') > 1:
                                        parts = cell_normalized.split('.')
                                        integer_part = ''.join(parts[:-1])
                                        decimal_part = parts[-1]
                                        cell_normalized = f"{integer_part}.{decimal_part}"
                                    
                                    num = float(cell_normalized)
                                    formatted = f"{num:.2f}"
                                    fila_limpia.append(formatted)
                                    
                                except (ValueError, AttributeError):
                                    fila_limpia.append("0.00")
            
            # Solo agregar si la fila tiene algún contenido
            if any(cell != "" for cell in fila_limpia):
                datos_extraidos.append(fila_limpia)
        
        wb.close()
        
        print(f"📊 Extraídos {len(datos_extraidos)} filas de {nombre_archivo} (hoja {hoja_mes}, desde fila {fila_inicio})")
        
        # Mostrar ejemplos detallados para debug
        if datos_extraidos and len(datos_extraidos) > 0:
            primera_fila = datos_extraidos[0]
            print(f"   🔍 Verificación de formato (primer registro):")
            
            if len(primera_fila) > 3:
                # Columna D (Nombre)
                nombre_original = primera_fila[3]
                nombre_sin_tildes = eliminar_tildes(nombre_original) if nombre_original else ""
                print(f"     Col D (Nombre original): '{nombre_original}'")
                print(f"     Col D (Nombre sin tildes): '{nombre_sin_tildes}'")
                
            if len(primera_fila) > 7:
                # Columna H (Repartición)
                reparticion_original = primera_fila[7]
                reparticion_sin_tildes = eliminar_tildes(reparticion_original) if reparticion_original else ""
                print(f"     Col H (Repartición original): '{reparticion_original}'")
                print(f"     Col H (Repartición sin tildes): '{reparticion_sin_tildes}'")
            
            # Mostrar ejemplo de conversión
            if len(primera_fila) > 3 and "á" in primera_fila[3] or "é" in primera_fila[3] or "í" in primera_fila[3] or "ó" in primera_fila[3] or "ú" in primera_fila[3]:
                print(f"     ✅ Ejemplo de conversión de tildes aplicado correctamente")
        
        return datos_extraidos
        
    except Exception as e:
        print(f"❌ Error extrayendo datos de {nombre_archivo}: {e}")
        import traceback
        traceback.print_exc()
        return []


def sanitizar_libro_remover_filtros(fh):
    """
    Intenta reparar un archivo .xlsx o .xlsm eliminando nodos <autoFilter>
    y otros nodos problemáticos. Devuelve BytesIO con el archivo reempaquetado
    o None si no pudo repararlo.
    """
    try:
        tmpdir = tempfile.mkdtemp()
        ruta_entrada = os.path.join(tmpdir, "input_file")
        with open(ruta_entrada, "wb") as f:
            f.write(fh.getvalue())

        ruta_extraida = os.path.join(tmpdir, "extracted")
        os.makedirs(ruta_extraida, exist_ok=True)

        with zipfile.ZipFile(ruta_entrada, "r") as zin:
            zin.extractall(ruta_extraida)

        ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        worksheets_dir = os.path.join(ruta_extraida, "xl", "worksheets")

        if os.path.isdir(worksheets_dir):
            for nombre_arch in os.listdir(worksheets_dir):
                if not nombre_arch.lower().endswith(".xml"):
                    continue
                ruta = os.path.join(worksheets_dir, nombre_arch)
                try:
                    tree = ET.parse(ruta)
                    root = tree.getroot()

                    modificado = False
                    for auto in root.findall("main:autoFilter", ns):
                        root.remove(auto)
                        modificado = True

                    for ext in root.findall("main:extLst", ns):
                        root.remove(ext)
                        modificado = True

                    for fcol in root.findall(".//main:filterColumn", ns):
                        for filt in fcol.findall("main:filters", ns):
                            fcol.remove(filt)
                            modificado = True

                    if modificado:
                        tree.write(ruta, encoding="utf-8", xml_declaration=True)
                except ET.ParseError:
                    # Si no se puede parsear una hoja, seguimos con las demás
                    pass

        ruta_salida = os.path.join(tmpdir, "output_file.xlsx")
        with zipfile.ZipFile(ruta_salida, "w", zipfile.ZIP_DEFLATED) as zout:
            for root_dir, dirs, files in os.walk(ruta_extraida):
                for file in files:
                    fullpath = os.path.join(root_dir, file)
                    arcname = os.path.relpath(fullpath, ruta_extraida)
                    zout.write(fullpath, arcname)

        with open(ruta_salida, "rb") as f:
            datos = f.read()

        shutil.rmtree(tmpdir)
        return BytesIO(datos)

    except Exception as e:
        try:
            shutil.rmtree(tmpdir)
        except Exception:
            pass
        print(f"❌ sanitizar_libro_remover_filtros error: {e}")
        return None