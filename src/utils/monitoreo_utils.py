"""
Utilidades del bot de monitoreo de liquidaciones.

Contiene:
  - CONFIG / constantes (espejo del Apps Script)
  - Lógica de comparación (normal y caja)
  - Generadores de adjuntos: XLSX de cambios y CSVs
  - Helpers de nombres/periodos
"""

import os
import re
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from utils.config_drive import FOLDER_REPARTICIONES_ID, FOLDER_SERVICES_ID
from utils.common_utils import ejecutar_con_reintentos_sheets


# ---------------------------------------------------------------------------
# Configuración (espejo del CONFIG del Apps Script)
# ---------------------------------------------------------------------------

CONFIG = {
    "CARPETA_REPARTICIONES_ID": FOLDER_REPARTICIONES_ID,
    "CARPETA_SERVICES_ID": FOLDER_SERVICES_ID,
    "CARPETA_SNAPSHOTS": "_snapshots_liquidaciones",
    "NOMBRE_REGISTRO": "_registro_agentes",

    "COL_INICIO": 1,
    "COL_FIN": 24,
    "COL_DNI": 2,

    "FILA_INICIO_DEFAULT": 4,
    "FILA_INICIO_CAJA": 5,
}

HOJAS_ORDEN = [
    "01", "02", "03", "04", "05", "06", "1° sac",
    "07", "08", "09", "10", "11", "12", "2° sac",
]

# Columnas numéricas (0-based dentro del rango A..X): cols 9-24 → offset 8-23
COLS_NUMERICAS = set(range(8, 24))

NOMBRES_COLUMNAS = [
    "cuil", "dni", "tipo doc", "nombre y apellido", "cod. liq.",
    "sit. revista", "estado afil.", "reparticion", "aporte personal",
    "adherente sec.", "fondo vol.", "hijo menor de 35", "menor a cargo",
    "cred. asist.", "sueldo sin desc.", "sueldo con desc.",
    "reaj. aporte pers.", "reaj. adh. sec.", "reaj. fv",
    "reaj. hijo menor", "reaj. menor cargo", "reaj. cred. asist.",
    "aporte patronal", "reaj. ap. patronal",
]

ENCABEZADOS_EXCEL = [
    "1-cuil", "2-dni", "3-tipo doc", "4-nombre y apellido", "5-cod. liq.",
    "6-sit. revista", "7-estado afil.", "8-reparticion", "9-aporte personal",
    "10-adherente sec.", "11-fondo vol.", "12-hijo menor de 35", "13-menor a cargo",
    "14-cred. asist.", "15-sueldo sin desc.", "16-sueldo con desc.",
    "17-reaj. aporte pers.", "18-reaj. adh. sec.", "19-reaj. fv",
    "20-reaj. hijo menor", "21-reaj. menor cargo", "22-reaj. cred. asist.",
    "23-aporte patronal", "24-reaj. ap. patronal",
]

# Columna de sueldo sin descuentos (0-based) para separar complementarias
COL_SUELDO_SIN_DESC = 14


# ---------------------------------------------------------------------------
# Helpers numéricos / texto
# ---------------------------------------------------------------------------

def parse_numero(val):
    if val is None or val == "":
        return 0.0
    if isinstance(val, (int, float)):
        return 0.0 if (isinstance(val, float) and (val != val)) else float(val)
    s = str(val).strip()
    if not s:
        return 0.0
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def formatear_importe(val):
    return f"{parse_numero(val):.2f}"


def normalizar_texto(s):
    import unicodedata
    t = str(s or "").upper()
    t = unicodedata.normalize("NFD", t)
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return " ".join(t.split())


def limpiar_numero(s):
    return re.sub(r"[^0-9]", "", str(s or ""))


def normalizar_cuil(val, col_idx):
    """
    Normaliza CUIL (col 0) y DNI (col 1) dejando solo los dígitos, para que
    un cambio de formato (guiones, puntos, espacios) no se reporte como
    "modificado" siendo el mismo número.

    Antes esto solo sacaba "-" del CUIL y ".-espacio" del DNI con una
    lista fija de símbolos. Formatos como "27 11312527 1" (espacios en el
    CUIL, sin guiones) no quedaban cubiertos y se comparaban tal cual,
    generando falsos "modificado". Como CUIL y DNI son identificadores
    puramente numéricos, lo correcto es sacar cualquier caracter que no
    sea dígito (0-9), sea cual sea el símbolo usado.

    Ejemplos:
      "27-11312527-1" / "27 11312527 1" / "27113125271" -> "27113125271"
      "38-261-080"    / "38.261.080"    / "38261080"    -> "38261080"
    """
    if col_idx in (0, 1):
        return re.sub(r"[^0-9]", "", val)
    return val


_RE_SEÑAL_MOJIBAKE = re.compile(r"[Ã\x80-\x9f]")
_RE_ENMASCARAR_COMPARACION = re.compile(r"[ÃÁÉÍÓÚÑãáéíóúñ\x80-\x9f]")


def _hay_señal_mojibake(s):
    return bool(_RE_SEÑAL_MOJIBAKE.search(s))


def _enmascarar_para_comparar(s):
    """
    Reduce a un mismo comodín ('#') tanto los restos de mojibake sin
    resolver (una "Ã" suelta, o un carácter de control 0x80-0x9f que quedó
    de un byte roto) como las vocales/ñ acentuadas — que son justamente lo
    que ese mojibake reconstruye cuando el byte SÍ está completo del otro
    lado. Se usa solo como comparación de respaldo (ver _valores_difieren),
    para reconocer que "NICOLÁS" (recuperado del lado que tenía el byte
    completo) y "NICOLÃS" (lado que ya lo había perdido en una corrupción
    previa a este pipeline) son el mismo dato de origen, sin tener que
    adivinar cuál letra falta.
    """
    return _RE_ENMASCARAR_COMPARACION.sub("#", s)


def _valores_difieren(va, vs, col_idx):
    """
    Compara dos valores de una misma columna para decidir si hay un cambio
    real que reportar como "modificado".

    Para columnas NUMÉRICAS (COLS_NUMERICAS, ej. aportes/reajustes) compara
    por VALOR NUMÉRICO (parse_numero, redondeado a 2 decimales) en vez de
    por string. El archivo "actual" y el snapshot representan el mismo cero
    (o el mismo importe) con strings distintos según de dónde vino la celda:
    "", "-", "0", "0.0", "0,00" son todos 0,00 en la práctica, pero como
    string son valores distintos. Compararlos como texto generaba falsos
    "modificado" con "Valor anterior" = "Valor nuevo" = "0,00" en el reporte
    (mismo número, formato de origen distinto), sin ningún cambio real que
    mostrarle al usuario.

    Para columnas NO numéricas (texto) se compara por string, con UNA
    excepción: si hay diferencia Y alguno de los dos lados todavía tiene
    señal de mojibake sin resolver (ver _reparar_texto_corrupto — esto pasa
    cuando un lado ya perdió el byte en una corrupción vieja, anterior a
    este pipeline, y el otro no), se prueba una segunda comparación
    "enmascarada" (_enmascarar_para_comparar) antes de decidir. Si al
    enmascarar ambos lados quedan iguales, NO se reporta como cambio real:
    es el mismo dato de origen corrupto, solo que un lado pudo reconstruir
    la letra real y el otro no. Este control es angosto a propósito (solo
    dispara si hay señal de mojibake) para no tapar ediciones reales de
    acentos en texto que nunca estuvo corrupto.
    """
    if col_idx in COLS_NUMERICAS:
        return round(parse_numero(va), 2) != round(parse_numero(vs), 2)
    if va == vs:
        return False
    if isinstance(va, str) and isinstance(vs, str) and (_hay_señal_mojibake(va) or _hay_señal_mojibake(vs)):
        if _enmascarar_para_comparar(va) == _enmascarar_para_comparar(vs):
            return False
    return True


# ---------------------------------------------------------------------------
# Lectura de rango (recibe lista de filas ya leída)
# ---------------------------------------------------------------------------

def leer_rango(datos):
    """
    Filtra filas donde el DNI (col offset 1) esté vacío, sea '-' o '0'.
    Recibe lista de listas tal como vienen de openpyxl values_only.
    """
    col_dni = CONFIG["COL_DNI"] - CONFIG["COL_INICIO"]   # 0-based
    return [
        fila for fila in datos
        if str(fila[col_dni] if col_dni < len(fila) else "").strip() not in ("", "-", "0")
    ]


def _normalizar_nombre_hoja(s):
    """Tolerante a variantes de nombre de hoja: mayúsc/minúsc, º/°, espacios."""
    return str(s).strip().lower().replace("º", "°").replace("  ", " ")


_RE_ESCAPE_LITERAL = re.compile(r"_x([0-9A-Fa-f]{4})_")


def _reparar_texto_corrupto(s):
    """
    Repara dos formas de corrupción de codificación detectadas en datos
    reales, mezcladas dentro de las mismas columnas de texto (ej.
    "8-reparticion", "4-nombre y apellido"). El mismo dato corrupto de
    origen llega representado de forma distinta según el camino de
    lectura: openpyxl directo del .xlsx actual, o Sheets API para el
    snapshot. Ejemplo real: "MUNICIPALIDAD DE COLONIA AYUÍ" corrupta.
      - snapshot (Sheets API):  ...AYUÃ\x8d   (carácter de control real)
      - actual (openpyxl):      ...AYUÃ_x008d_  (el escape como texto literal)

    Paso 1 — Escape literal de Excel/OOXML sin resolver:
    un carácter inválido para XML 1.0 se guarda como texto "_xHHHH_" en vez
    del carácter real. openpyxl normalmente ya lo resuelve solo; el .xlsx
    de origen (generado por otro sistema, antes de llegar a este pipeline)
    a veces deja la secuencia visible como texto. Se deshace ANTES del
    paso 2: hasta no tener el carácter de control real, el mojibake de
    abajo no es reconocible como tal.

    Paso 2 — Mojibake real (doble-decodificación):
    bytes UTF-8 válidos que en algún momento ANTERIOR a este pipeline se
    decodificaron con la codificación equivocada. En los datos reales
    aparecieron dos variantes históricas distintas -- Latin-1 y
    Windows-1252 -- así que se prueban ambas y se usa la primera que
    efectivamente reconstruye UTF-8 válido. Solo se actúa cuando:
      (a) el string tiene una señal de corrupción (un carácter de control
          fuera de \\t\\n\\r, típico del rango C1 0x80-0x9F que deja este
          tipo de mojibake), y
      (b) el re-decode da un resultado DISTINTO y además imprimible.
    Sin estas dos condiciones se deja el valor tal cual: intentar esto
    sobre texto que ya está bien puede introducir corrupción nueva donde
    no la había.

    Límite conocido (no resoluble automáticamente): si el dato ya perdió
    un byte en una corrupción previa más vieja (ej. snapshot con
    "NICOLÃS" en vez de "NICOLÃ\x81S" -- falta directamente el byte, no
    quedó ni el carácter de control), no hay bytes que recuperar y la
    función lo deja como está. Ese caso puntual requiere corrección manual
    una vez; después de esa corrección (o cuando el archivo actual, ya
    reparado acá, se vuelque al snapshot en la próxima corrida) se
    autocorrige solo.
    """
    if "_x" in s:
        s = _RE_ESCAPE_LITERAL.sub(lambda m: chr(int(m.group(1), 16)), s)

    tiene_señal = any(
        (0x80 <= ord(c) <= 0x9F) or (ord(c) < 0x20 and c not in "\t\n\r")
        for c in s
    )
    if tiene_señal:
        for codificacion in ("latin-1", "cp1252"):
            try:
                candidato = s.encode(codificacion).decode("utf-8")
            except (UnicodeDecodeError, UnicodeEncodeError):
                continue
            if candidato != s and candidato.isprintable():
                s = candidato
                break

    return s


def _normalizar_valor_celda(val):
    """
    Normaliza un valor recién leído de una hoja .xlsx con openpyxl.

    El snapshot es un Google Sheet exportado a .xlsx. Google Sheets guarda
    TODOS los números como float de doble precisión, así que un CUIL, DNI,
    tipo de doc. o código de liquidación que en el archivo "actual"
    original llega como int (ej. 27100734805) vuelve del snapshot como
    27100734805.0. openpyxl entrega ambos casos tal cual los encuentra.

    Si no se corrige acá, todo lo que compara/agrupa/muestra usando
    str(valor) — identificación de agentes (_indexar / obtener_id_agente),
    detección de modificaciones (comparar_hojas_normal/caja), el Excel de
    cambios y el CSV — trata "27100734805.0" y "27100734805" como valores
    DISTINTOS. Consecuencia real (no solo estética): la misma persona
    termina detectada como "eliminada" (con el valor que vino del
    snapshot, con .0) y como "nueva" (con el valor del archivo actual,
    sin .0) en la misma corrida, en vez de reconocerse como el mismo
    registro sin cambios.

    Esta función deja los floats que representan un entero (ej. 8653.0,
    27100734805.0) como int, igual a como ya llegan naturalmente desde el
    archivo actual. Los floats con decimales reales (importes) no se
    tocan — igual se siguen formateando después con parse_numero().

    Para strings, aplica además _reparar_texto_corrupto() (ver docstring)
    para que ambos lados de cada comparación (archivo actual y snapshot)
    queden en la misma representación, y para recuperar el texto real
    cuando los bytes originales todavía están disponibles.
    """
    if isinstance(val, float) and val.is_integer():
        return int(val)
    if isinstance(val, str) and val:
        return _reparar_texto_corrupto(val)
    return val


def leer_hoja_xlsx(fh, nombre_hoja, fila_inicio, col_fin=None):
    """
    Lee una hoja específica de un archivo .xlsx (o de un Google Sheet
    exportado a .xlsx, que es cómo llega el snapshot) usando openpyxl.

    No necesita convertir nada a Google Sheets ni usar la Sheets API —
    tanto el archivo actual como el snapshot exportado se leen igual,
    con el mismo código que ya usan fv_drive_bot.py y reporte_anual_bot.py.

    Devuelve una lista de filas (listas), desde `fila_inicio` hasta que
    encuentra una fila vacía o con '-' en la columna A. Si la hoja no
    existe (por nombre, tolerando variantes º/°), devuelve [].

    Cada valor pasa por _normalizar_valor_celda() para que el archivo
    actual y el snapshot (Google Sheet) queden en el mismo formato — ver
    el docstring de esa función para el motivo.
    """
    col_fin = col_fin or CONFIG["COL_FIN"]
    fh.seek(0)
    wb = openpyxl.load_workbook(fh, data_only=True, read_only=True)
    try:
        objetivo = _normalizar_nombre_hoja(nombre_hoja)
        hoja_real = None
        for nombre in wb.sheetnames:
            if _normalizar_nombre_hoja(nombre) == objetivo:
                hoja_real = nombre
                break
        if hoja_real is None:
            return []

        ws = wb[hoja_real]
        filas = []
        for row in ws.iter_rows(min_row=fila_inicio, max_col=col_fin, values_only=True):
            primera = row[0] if row else None
            if primera is None or str(primera).strip() in ("", "-"):
                break
            filas.append([_normalizar_valor_celda(v) for v in row])
        return filas
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Actualización del snapshot (Sheets API, "in-place")
# ---------------------------------------------------------------------------
#
# El snapshot es un Google Sheet que ya existe (lo crea snapshot_bot.py con
# OAuth). monitoreo_bot, que corre con Service Account, NO puede recrear ese
# archivo — pero SÍ puede sobrescribir el CONTENIDO de sus pestañas, que es
# una operación de edición, no de creación. Con esto evitamos también el
# viejo patrón de "borrar y volver a subir una copia nueva" en cada corrida.

def asegurar_hoja_snapshot(sheets_svc, spreadsheet_id, nombre_hoja):
    """
    Verifica que exista una pestaña con ese nombre dentro del snapshot; si
    no existe, la crea (agregar una pestaña a un Sheet existente es edición
    de contenido, no creación de archivo — la Service Account puede hacerlo).
    """
    meta = ejecutar_con_reintentos_sheets(
        sheets_svc.spreadsheets().get(spreadsheetId=spreadsheet_id),
        f"leer metadata del snapshot {spreadsheet_id} (asegurar_hoja_snapshot)",
    )
    hojas = [s["properties"]["title"] for s in meta.get("sheets", [])]
    if nombre_hoja in hojas:
        return
    ejecutar_con_reintentos_sheets(
        sheets_svc.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": [{"addSheet": {"properties": {"title": nombre_hoja}}}]},
        ),
        f"crear pestaña '{nombre_hoja}' en snapshot {spreadsheet_id}",
    )


def actualizar_snapshot_hoja(sheets_svc, spreadsheet_id, nombre_hoja, filas, fila_inicio=1, col_fin=None):
    """
    Sobrescribe el contenido de una pestaña del snapshot con `filas`
    (lista de listas). Primero limpia el rango completo y después escribe
    los valores nuevos, para no dejar residuos de filas que ya no existen.

    IMPORTANTE: `filas` son SOLO datos (sin las filas de encabezado/metadata
    que tiene el .xlsx original en las filas 1..fila_inicio-1). Si se
    escribieran siempre a partir de A1, el snapshot quedaría con los datos
    corridos hacia arriba respecto de cómo se creó originalmente (copia
    cruda del .xlsx vía snapshot_bot.py, que sí conserva esas filas de
    metadata). Como leer_hoja_xlsx() siempre lee el snapshot arrancando en
    `fila_inicio` (el mismo que usa para el archivo real), hay que escribir
    también a partir de esa fila para no correr el registro: si no, en la
    próxima corrida los primeros (fila_inicio - 1) agentes reales dejan de
    estar donde el lector los espera y se reportan como "nuevos" en cada
    ejecución.
    """
    col_fin = col_fin or CONFIG["COL_FIN"]
    ultima_col = _letra_columna(col_fin)

    asegurar_hoja_snapshot(sheets_svc, spreadsheet_id, nombre_hoja)

    ejecutar_con_reintentos_sheets(
        sheets_svc.spreadsheets().values().clear(
            spreadsheetId=spreadsheet_id,
            range=f"'{nombre_hoja}'!A1:{ultima_col}200000",
            body={},
        ),
        f"limpiar pestaña '{nombre_hoja}' del snapshot {spreadsheet_id}",
    )

    if filas:
        ejecutar_con_reintentos_sheets(
            sheets_svc.spreadsheets().values().update(
                spreadsheetId=spreadsheet_id,
                range=f"'{nombre_hoja}'!A{fila_inicio}",
                valueInputOption="RAW",
                body={"values": filas},
            ),
            f"escribir pestaña '{nombre_hoja}' del snapshot {spreadsheet_id}",
        )


def _letra_columna(n):
    letra = ""
    while n > 0:
        n, resto = divmod(n - 1, 26)
        letra = chr(65 + resto) + letra
    return letra


# ---------------------------------------------------------------------------
# Lectura del snapshot (Sheets API, sin exportar a .xlsx)
# ---------------------------------------------------------------------------
#
# drive.files().export tiene un límite de tamaño del lado de Google
# ("This file is too large to be exported" / exportSizeLimitExceeded) que
# NO es transitorio: no hay reintento que lo resuelva. Con reparticiones
# grandes (ej. MUNICIPIO PARANA, .xlsx original de ~30MB) el export del
# snapshot lo pisa siempre, dejando esa repartición sin poder compararse
# nunca. La Sheets API no tiene ese límite porque no arma un binario con
# estilos: solo devuelve las celdas del rango pedido. Estas funciones leen
# el snapshot por ese camino, como reemplazo de exportar + leer con
# openpyxl (leer_hoja_xlsx), manteniendo el mismo formato de salida
# (lista de filas, valores normalizados) para no tener que tocar
# comparar_hojas_normal/caja.

def obtener_titulos_hojas_snapshot(sheets_svc, spreadsheet_id):
    """
    Devuelve {nombre_normalizado: título_real} para todas las pestañas del
    snapshot. Se pide UNA sola vez por archivo (no por hoja) y se reusa en
    cada llamada a leer_hojas_snapshot_batch(), para no multiplicar por 14 la
    cantidad de lecturas de metadata contra la Sheets API.
    """
    meta = ejecutar_con_reintentos_sheets(
        sheets_svc.spreadsheets().get(spreadsheetId=spreadsheet_id),
        f"leer metadata del snapshot {spreadsheet_id}",
    )
    return {
        _normalizar_nombre_hoja(s["properties"]["title"]): s["properties"]["title"]
        for s in meta.get("sheets", [])
    }


def leer_hojas_snapshot_batch(sheets_svc, spreadsheet_id, nombres_hojas, fila_inicio, titulos_hojas, col_fin=None):
    """
    Lee TODAS las pestañas de `nombres_hojas` del snapshot en UNA sola
    llamada a la Sheets API (values().batchGet), en vez de una llamada
    values().get() por pestaña (versión anterior: leer_hoja_snapshot).

    Con las ~14 hojas de HOJAS_ORDEN, esto baja el consumo de la cuota de
    lectura (60 req/min/usuario) ~14x por archivo — la causa real de los
    "Quota exceeded for quota metric 'Read requests'" vistos en producción.

    Misma semántica de salida POR HOJA que la función anterior: lista de
    filas desde fila_inicio hasta la primera fila vacía o con '-' en la
    columna A, con cada valor pasado por _normalizar_valor_celda(). Si una
    pestaña no existe en el snapshot (tolerante a variantes º/°, vía
    `titulos_hojas`), queda como [] en el resultado.

    A DIFERENCIA de la función anterior, si la llamada a la API falla (p.ej.
    cuota agotada tras los reintentos de ejecutar_con_reintentos_sheets), NO
    se traga el error devolviendo listas vacías — se propaga la excepción.
    Devolver [] en silencio hacía que el caller comparara el archivo actual
    contra "snapshot vacío", generando falsos positivos de "eliminado" en
    TODOS los agentes de esa pestaña. El caller (_procesar_archivo_impl en
    monitoreo_bot.py) atrapa esta excepción y omite la comparación de TODO
    el archivo esa corrida, sin tocar el snapshot — mismo patrón ya usado
    para fallos de obtener_snapshot_de_archivo() y obtener_titulos_hojas_snapshot().

    Devuelve {nombre_hoja: [filas]}.
    """
    col_fin = col_fin or CONFIG["COL_FIN"]
    ultima_col = _letra_columna(col_fin)

    nombres_validos = []
    rangos = []
    for nombre_hoja in nombres_hojas:
        titulo_real = titulos_hojas.get(_normalizar_nombre_hoja(nombre_hoja))
        if titulo_real is None:
            continue
        rangos.append(f"'{titulo_real}'!A{fila_inicio}:{ultima_col}")
        nombres_validos.append(nombre_hoja)

    resultado = {nombre_hoja: [] for nombre_hoja in nombres_hojas}
    if not rangos:
        return resultado

    resp = ejecutar_con_reintentos_sheets(
        sheets_svc.spreadsheets().values().batchGet(
            spreadsheetId=spreadsheet_id,
            ranges=rangos,
            valueRenderOption="UNFORMATTED_VALUE",
        ),
        f"leer {len(rangos)} pestaña(s) del snapshot {spreadsheet_id} (batchGet)",
    )

    # batchGet devuelve valueRanges en el MISMO orden que se pidieron los
    # ranges (garantizado por la API) — por eso alcanza con zip() en vez de
    # tener que parsear el string de range devuelto para reidentificar la
    # hoja (sería frágil con títulos que contienen comillas o "!").
    for nombre_hoja, value_range in zip(nombres_validos, resp.get("valueRanges", [])):
        filas = []
        for row in value_range.get("values", []):
            primera = row[0] if row else None
            if primera is None or str(primera).strip() in ("", "-"):
                break
            fila_norm = [_normalizar_valor_celda(v) for v in row]
            # Sheets recorta las celdas vacías al final de cada fila: hay que
            # completar hasta col_fin para no correr los offsets fijos que usa
            # el resto del código (CUIL, DNI, columnas numéricas, etc.)
            if len(fila_norm) < col_fin:
                fila_norm = fila_norm + [None] * (col_fin - len(fila_norm))
            filas.append(fila_norm)
        resultado[nombre_hoja] = filas

    return resultado


# ---------------------------------------------------------------------------
# Indexado por ID de agente
# ---------------------------------------------------------------------------

def _indexar(filas, hoja_registro, get_id_fn):
    col_dni = CONFIG["COL_DNI"] - CONFIG["COL_INICIO"]
    mapa = {}
    for fila in filas:
        cuil = str(fila[0] if len(fila) > 0 else "").strip()
        dni = str(fila[col_dni] if len(fila) > col_dni else "").strip()
        nombre = str(fila[3] if len(fila) > 3 else "").strip()
        if not cuil and not dni and not nombre:
            continue
        aid = get_id_fn(cuil, dni, nombre, hoja_registro) if hoja_registro else (dni or cuil)
        mapa.setdefault(aid, []).append(fila)
    return mapa


# ---------------------------------------------------------------------------
# Comparación modo normal
# ---------------------------------------------------------------------------

def _reordenar_ordinaria_primero(filas):
    """
    Reordena las filas de UN MISMO agente (mismo aid, ya agrupadas por
    _indexar) para que la fila "ordinaria" quede primera y el resto
    (complementarias) mantenga su orden relativo. Con una sola fila no
    hace nada.

    Usa el mismo criterio que separar_complementarias_agrupado (columna
    "sueldo sin desc.", COL_SUELDO_SIN_DESC): la fila con mayor sueldo es
    la ordinaria, las demás son complementarias.

    Por qué hace falta: comparar_hojas_normal empareja filas de un mismo
    agente por posición (filas_act[i] contra filas_sn[i]). Si un agente
    tiene una liquidación ordinaria + una o más complementarias, y el
    orden en que aparecen esas filas en la planilla cambia entre el
    archivo actual y el snapshot (ej. antes la complementaria aparecía
    primero y ahora aparece después), sin este reordenamiento se termina
    comparando la ordinaria contra la complementaria (o viceversa) solo
    por una diferencia de orden -- generando "modificado" en TODAS las
    columnas de esa fila, aunque ninguna haya cambiado realmente. Al fijar
    la ordinaria siempre en la posición 0 de ambos lados, se empareja
    ordinaria-con-ordinaria y complementaria(s)-con-complementaria(s).

    Nota: con 2+ complementarias para un mismo agente, estas se siguen
    emparejando entre sí por orden de aparición (no hay forma de
    distinguirlas individualmente con los datos disponibles), así que ese
    caso más raro puede seguir generando falsos "modificado" si además
    cambia el orden ENTRE complementarias. El caso común (una ordinaria +
    a lo sumo una complementaria) queda cubierto.
    """
    if len(filas) <= 1:
        return filas
    idx_max = max(
        range(len(filas)),
        key=lambda i: parse_numero(filas[i][COL_SUELDO_SIN_DESC] if len(filas[i]) > COL_SUELDO_SIN_DESC else 0),
    )
    resto = [f for i, f in enumerate(filas) if i != idx_max]
    return [filas[idx_max]] + resto


def comparar_hojas_normal(datos_actual, datos_snap, hoja_registro):
    from utils.registro_utils import obtener_id_agente
    cambios = []
    cant_cols = CONFIG["COL_FIN"] - CONFIG["COL_INICIO"] + 1
    col_dni = CONFIG["COL_DNI"] - CONFIG["COL_INICIO"]

    mapa_snap = _indexar(datos_snap, hoja_registro, obtener_id_agente)
    mapa_act = _indexar(datos_actual, hoja_registro, obtener_id_agente)

    # Eliminados completamente
    for aid, filas in mapa_snap.items():
        if aid not in mapa_act:
            f = filas[0]
            dni = str(f[col_dni] if len(f) > col_dni else "").strip()
            nombre = f[3] if len(f) > 3 else "(sin nombre)"
            cambios.append({"tipo": "eliminado", "dni": dni, "nombre": nombre, "fila": f})

    for aid, filas_act in mapa_act.items():
        fref = filas_act[0]
        dni = str(fref[col_dni] if len(fref) > col_dni else "").strip()
        nombre = fref[3] if len(fref) > 3 else "(sin nombre)"

        if aid not in mapa_snap:
            for f in filas_act:
                cambios.append({"tipo": "nuevo", "dni": dni, "nombre": nombre, "fila": f})
            continue

        filas_sn = mapa_snap[aid]

        # Fija la fila "ordinaria" de cada lado en la posición 0 antes de
        # emparejar por índice, para no comparar ordinaria contra
        # complementaria solo por un cambio de orden en la planilla (ver
        # docstring de _reordenar_ordinaria_primero).
        filas_act = _reordenar_ordinaria_primero(filas_act)
        filas_sn = _reordenar_ordinaria_primero(filas_sn)

        # Filas eliminadas (había más antes)
        for i in range(len(filas_act), len(filas_sn)):
            cambios.append({"tipo": "eliminado", "dni": dni, "nombre": nombre, "fila": filas_sn[i]})

        # Filas nuevas (hay más ahora)
        for i in range(len(filas_sn), len(filas_act)):
            cambios.append({"tipo": "nuevo", "dni": dni, "nombre": nombre, "fila": filas_act[i]})

        # Comparar fila a fila
        for i in range(min(len(filas_act), len(filas_sn))):
            fa, fs = filas_act[i], filas_sn[i]
            for c in range(cant_cols):
                # OJO: guardar explícitamente contra "fa[c] is not None", no
                # solo contra índice fuera de rango. openpyxl (archivo actual)
                # deja una celda vacía como Python None, y str(None) da el
                # texto "None" (truthy) en vez de "" -- eso hacía que el
                # reporte de "modificado" mostrara literalmente "None" en
                # vez de "(vacío)" para campos de texto vacíos.
                va = normalizar_cuil(str(fa[c]).strip() if (len(fa) > c and fa[c] is not None) else "", c)
                vs = normalizar_cuil(str(fs[c]).strip() if (len(fs) > c and fs[c] is not None) else "", c)
                if _valores_difieren(va, vs, c):
                    cambios.append({
                        "tipo": "modificado",
                        "id": aid,
                        "dni": dni,
                        "nombre": nombre,
                        "columna": NOMBRES_COLUMNAS[c] if c < len(NOMBRES_COLUMNAS) else f"col{c+1}",
                        "anterior": vs or "(vacío)",
                        "actual": va or "(vacío)",
                        "es_no_numerico": c not in COLS_NUMERICAS,
                        "fila": fa,
                    })


    return {"cambios": cambios, "mapa_actual": mapa_act}


# ---------------------------------------------------------------------------
# Comparación modo caja
# ---------------------------------------------------------------------------

# Columna "6-sit. revista" (0-based dentro del rango A..X)
COL_SIT_REVISTA = 5



# Diccionario explícito de variantes conocidas -> concepto canónico, para
# "06-sit. revista" en archivos "Caja". Reemplaza al heurístico anterior
# de "empieza con JUB/PEN", que tenía dos riesgos:
#   - Falso positivo: un valor real que arranca con esas letras sin ser
#     jubilado/pensionado (ej. "PENDIENTE") quedaba mal agrupado.
#   - Falso negativo (si se sacaba el prefijo sin reemplazo): variantes de
#     escritura de un mismo concepto que no coinciden como texto exacto
#     ("JUB." vs "JUBILADO") dejaban de agruparse entre sí.
#
# Universo relevado sobre datos reales (unificado mensual, filtrado a
# reparticiones de Caja/Jubilaciones). Claves ya pasadas por
# normalizar_texto() + remoción de puntos + .upper() (mismo preprocesamiento
# que se les aplica antes de la búsqueda en el dict — ver
# _normalizar_situacion_revista). El matching es case-insensitive, así que
# alcanza con UNA clave en mayúsculas por variante: no hace falta cargar
# "Jubilado" y "JUBILADO" por separado.
CONCEPTOS_SIT_REVISTA = {
    # --- JUBILADO (incluye PASIVO y RETIRADO, confirmado con el usuario) ---
    "JUBILADO": "JUBILADO",
    "JUBILACION": "JUBILADO",
    "JUBIL": "JUBILADO",
    "JUBLADO": "JUBILADO",                          # typo detectado en datos reales
    "JUBILADO - ( 82% )": "JUBILADO",
    "JUBILACION ED ADA - (82%)": "JUBILADO",
    "JUBILACION EXTR (82%)": "JUBILADO",
    "JUBILACION ORDINARIA (85%)": "JUBILADO",
    "PASIVO": "JUBILADO",
    "RETIRADO": "JUBILADO",
    "JUBILADOS": "JUBILADO",

    # --- PENSIONADO (incluye MEDIA PENSION, confirmado con el usuario) ---
    "PENSIONADO": "PENSIONADO",
    "PENSION": "PENSIONADO",
    "PENSIONADA": "PENSIONADO",
    "PENSION - (75%)": "PENSIONADO",
    "MEDIA PENSION -  (75%)": "PENSIONADO",

    # --- ACTIVO (personal de planta de la propia caja, no jubilado/pensionado) ---
    "ACTIVO": "ACTIVO",
    "EMPL ACTIVO": "ACTIVO",
    "CONTRATADO": "ACTIVO",
    "EMPLEADO ADMINISTRATIVO": "ACTIVO",
    "MAESTRANZA": "ACTIVO",
    "FUNCIONARI": "ACTIVO",
    "FUNCIONARIO": "ACTIVO",
    "PTA PERMANENTE": "ACTIVO",
    "PERMANENTE": "ACTIVO",
    "CAT1": "ACTIVO",
}


def _normalizar_situacion_revista(val):
    """
    Normaliza el valor de "sit. revista" (columna 6) a un concepto
    canónico, para poder agrupar/emparejar las liquidaciones de un mismo
    agente por CONCEPTO en vez de por posición en la planilla.

    En modo "caja" un mismo agente (mismo DNI/CUIL, mismo aid) puede tener
    DOS o más liquidaciones legítimas y distintas dentro del mismo período
    (ej. una como JUBILADO y otra como PENSIONADO). El texto no está
    estandarizado entre reparticiones/meses ("JUBILADO", "Jubilado",
    "Jubil.", "PASIVO", "RETIRADO", "PENSIONADO", "Pension.", "MEDIA
    PENSION - (75%)", etc.), así que las variantes conocidas se resuelven
    contra CONCEPTOS_SIT_REVISTA (ver arriba) en vez de con un heurístico
    de prefijo — evita tanto falsos positivos (un valor no relacionado que
    por casualidad empieza con esas letras) como falsos negativos
    (variantes de escritura que antes no colapsaban al no matchear texto
    exacto). Cualquier valor NO listado en el diccionario (o vacío, agente
    sin este dato declarado) se deja normalizado tal cual (mayúsculas, sin
    tildes ni puntos) y actúa como su propio concepto — típicamente
    personal de planta de la caja (CONTRATADO, FUNCIONARIO, etc.), que no
    debe forzarse a jubilado/pensionado.

    Se mayusculiza ANTES de buscar en el diccionario (y también en el
    fallback) para que el match no dependa de cómo esté tipeado el valor
    en la planilla. Sin esto, "JUBILADO" y "Jubilado" son 2 claves
    distintas que habría que cargar por separado en CONCEPTOS_SIT_REVISTA
    solo por una diferencia de mayúsculas/minúsculas — con .upper() alcanza
    con una sola entrada por concepto, sea cual sea el case con el que
    venga escrito en el archivo de origen.
    """
    t = normalizar_texto(val).replace(".", "").strip().upper()
    return CONCEPTOS_SIT_REVISTA.get(t, t)


def _agrupar_por_concepto(filas):
    """
    Agrupa las filas de UN MISMO agente (mismo aid, ya agrupadas por
    _indexar) por concepto (_normalizar_situacion_revista: JUBILADO /
    PENSIONADO / otro), y dentro de cada concepto deja la fila "ordinaria"
    primero (_reordenar_ordinaria_primero) — mismo criterio de sueldo sin
    desc. que en modo normal.

    Por qué hace falta: en modo caja, comparar_hojas_caja emparejaba las
    filas de un agente por posición cruda dentro de su lista. Un agente
    JUBILADO + PENSIONADO ya son 2 filas por sí solas (sin que eso sea una
    complementaria); si además una de esas dos tiene su propia
    complementaria, son 3 filas. Emparejar por índice puro comparaba, por
    ejemplo, la liquidación de JUBILADO contra la de PENSIONADO (o la
    ordinaria contra la complementaria) apenas cambiaba el orden en que
    aparecen en la planilla entre el snapshot y el archivo actual —
    generando "modificado" en TODAS las columnas de esa fila sin que haya
    ningún cambio real. Agrupando primero por concepto y después por
    ordinaria/complementaria dentro de cada concepto, el emparejamiento
    deja de depender del orden de aparición.
    """
    grupos = {}
    for f in filas:
        concepto = _normalizar_situacion_revista(f[COL_SIT_REVISTA] if len(f) > COL_SIT_REVISTA else "")
        grupos.setdefault(concepto, []).append(f)
    return {concepto: _reordenar_ordinaria_primero(fs) for concepto, fs in grupos.items()}


def _revisar_conceptos_sin_mapear(grupos, dni, nombre):
    """
    Aviso SOLO para logs (nunca para el mail — ver comparar_hojas_caja y
    monitoreo_bot.py, que lo imprime aparte del armado del HTML).

    Detecta el único caso donde un valor de "sit. revista" sin mapear en
    CONCEPTOS_SIT_REVISTA puede arruinar el agrupamiento ordinaria/
    complementaria: cuando el agente tiene 2+ CONCEPTOS distintos en la
    misma corrida (ya pasó el filtro de "más de una fila" en
    comparar_hojas_caja) y al menos uno de esos conceptos no está en el
    diccionario — es decir, quedó como su propio concepto por el
    fallback de _normalizar_situacion_revista().

    Si el agente tiene un solo concepto (aunque ese concepto sea un
    fallback, ej. 2 filas "CONTRATADO"), NO hay aviso: ahí no hay
    ambigüedad, las filas ya agrupan correctamente entre sí.
    """
    if len(grupos) <= 1:
        return []
    avisos = []
    for concepto, filas in grupos.items():
        if concepto in CONCEPTOS_SIT_REVISTA:
            continue
        raw = filas[0][COL_SIT_REVISTA] if len(filas[0]) > COL_SIT_REVISTA else ""
        avisos.append(
            f"DNI {dni} ({nombre}): valor de sit. revista sin mapear "
            f"'{raw}' (normalizado: '{concepto}') conviviendo con otro(s) "
            f"concepto(s) en el mismo período — revisar si es variante de "
            f"JUBILADO/PENSIONADO en CONCEPTOS_SIT_REVISTA."
        )
    return avisos


def comparar_hojas_caja(datos_actual, datos_snap, hoja_registro):
    from utils.registro_utils import obtener_id_agente
    cambios = []
    cant_cols = CONFIG["COL_FIN"] - CONFIG["COL_INICIO"] + 1
    col_dni = CONFIG["COL_DNI"] - CONFIG["COL_INICIO"]

    grupos_act = _indexar(datos_actual, hoja_registro, obtener_id_agente)
    grupos_snap = _indexar(datos_snap, hoja_registro, obtener_id_agente)
    todos_ids = set(list(grupos_act.keys()) + list(grupos_snap.keys()))
    avisos_conceptos = []

    for aid in todos_ids:
        fa_list = grupos_act.get(aid, [])
        fs_list = grupos_snap.get(aid, [])
        fref = (fa_list or fs_list)[0]
        dni = str(fref[col_dni] if len(fref) > col_dni else "").strip()
        nombre = fref[3] if len(fref) > 3 else "(sin nombre)"

        # Agrupa por concepto (JUBILADO / PENSIONADO / otro) SOLO cuando
        # hace falta desambiguar -- es decir, cuando el agente tiene más
        # de una fila de algún lado (jubilado+pensionado, con o sin
        # complementaria). Ver docstring de _agrupar_por_concepto: evita
        # comparar JUBILADO contra PENSIONADO (o ordinaria contra
        # complementaria) solo por un cambio de orden.
        #
        # Cuando el agente tiene como máximo 1 fila de cada lado (el caso
        # más común: una sola liquidación, sin jubilado+pensionado) se
        # compara DIRECTO, sin pasar por el concepto. Si se agrupara
        # siempre por concepto, un cambio real de "sit. revista" entre
        # snapshot y actual (ej. Activo -> Baja) haría que la fila caiga
        # en dos "conceptos" distintos y se reporte como
        # eliminado+nuevo en vez de "modificado" en esa columna --
        # perdiendo el detalle del cambio para el caso más frecuente.
        if len(fa_list) <= 1 and len(fs_list) <= 1:
            conceptos_act = {"__unico__": fa_list}
            conceptos_snap = {"__unico__": fs_list}
        else:
            conceptos_act = _agrupar_por_concepto(fa_list)
            conceptos_snap = _agrupar_por_concepto(fs_list)
            avisos_conceptos.extend(_revisar_conceptos_sin_mapear(conceptos_act, dni, nombre))
        todos_conceptos = set(list(conceptos_act.keys()) + list(conceptos_snap.keys()))

        for concepto in todos_conceptos:
            filas_c_act = conceptos_act.get(concepto, [])
            filas_c_snap = conceptos_snap.get(concepto, [])

            for i in range(len(filas_c_act), len(filas_c_snap)):
                cambios.append({
                    "tipo": "eliminado",
                    "dni": dni,
                    "nombre": nombre,
                    "registro": i + 1,
                    "fila": filas_c_snap[i]
                })

            for i in range(len(filas_c_snap), len(filas_c_act)):
                cambios.append({
                    "tipo": "nuevo",
                    "dni": dni,
                    "nombre": nombre,
                    "registro": i + 1,
                    "fila": filas_c_act[i]
                })

            for i in range(min(len(filas_c_act), len(filas_c_snap))):
                fa, fs = filas_c_act[i], filas_c_snap[i]
                for c in range(cant_cols):
                    # Ver comentario equivalente en comparar_hojas_normal:
                    # guarda contra fa[c]/fs[c] is None para no convertir
                    # celdas vacías en el texto literal "None".
                    va = normalizar_cuil(str(fa[c]).strip() if (len(fa) > c and fa[c] is not None) else "", c)
                    vs = normalizar_cuil(str(fs[c]).strip() if (len(fs) > c and fs[c] is not None) else "", c)
                    if _valores_difieren(va, vs, c):
                        cambios.append({
                            "tipo": "modificado",
                            "id": aid,
                            "dni": dni,
                            "nombre": nombre,
                            "registro": i + 1,
                            "columna": NOMBRES_COLUMNAS[c] if c < len(NOMBRES_COLUMNAS) else f"col{c+1}",
                            "anterior": vs or "(vacío)",
                            "actual": va or "(vacío)",
                            "es_no_numerico": c not in COLS_NUMERICAS,
                            "fila": fa,
                        })

    return {"cambios": cambios, "mapa_actual": grupos_act, "avisos_conceptos": avisos_conceptos}


# ---------------------------------------------------------------------------
# Separar ordinarias / complementarias
# ---------------------------------------------------------------------------

def separar_complementarias_agrupado(mapa_agrupado):
    """
    mapa_agrupado: { id: [fila, ...] }
    Retorna: { "ordinarias": {id: fila}, "complementarias": {id: [fila, ...]} }

    Modo NORMAL: un agente tiene como máximo una liquidación "ordinaria"
    por período, así que "ordinarias" queda con UNA fila por id.
    """
    ordinarias = {}
    complementarias = {}

    for aid, filas in mapa_agrupado.items():
        if not filas:
            continue
        if len(filas) == 1:
            ordinarias[aid] = filas[0]
            continue
        idx_max = max(range(len(filas)), key=lambda i: parse_numero(filas[i][COL_SUELDO_SIN_DESC] if len(filas[i]) > COL_SUELDO_SIN_DESC else 0))
        ordinarias[aid] = filas[idx_max]
        comps = [f for i, f in enumerate(filas) if i != idx_max]
        if comps:
            complementarias[aid] = comps

    return {"ordinarias": ordinarias, "complementarias": complementarias}


def separar_complementarias_agrupado_caja(mapa_agrupado):
    """
    Equivalente a separar_complementarias_agrupado() pero para modo
    "caja". mapa_agrupado: { id: [fila, ...] }
    Retorna: { "ordinarias": {id: [fila, ...]}, "complementarias": {id: [fila, ...]} }

    Diferencia clave con el modo normal: en "caja" un mismo agente puede
    tener MÁS DE UNA liquidación ordinaria legítima en el mismo período
    (ej. JUBILADO + PENSIONADO — ver _normalizar_situacion_revista), no
    solo una. Por eso primero se agrupa por concepto
    (_agrupar_por_concepto) y de cada concepto se toma su propia fila de
    mayor sueldo como ordinaria; el resto de cada concepto son
    complementarias. Aplicar directamente separar_complementarias_agrupado()
    acá tomaría la liquidación de menor sueldo entre JUBILADO y PENSIONADO
    como si fuera una complementaria del otro concepto, cuando en realidad
    son dos liquidaciones ordinarias distintas.

    Por eso "ordinarias" queda con LISTA de filas por id (una por
    concepto), a diferencia del modo normal que tiene una sola fila por id.
    """
    ordinarias = {}
    complementarias = {}

    for aid, filas in mapa_agrupado.items():
        if not filas:
            continue
        grupos = _agrupar_por_concepto(filas)
        ords, comps = [], []
        for filas_concepto in grupos.values():
            if not filas_concepto:
                continue
            ords.append(filas_concepto[0])
            comps.extend(filas_concepto[1:])
        if ords:
            ordinarias[aid] = ords
        if comps:
            complementarias[aid] = comps

    return {"ordinarias": ordinarias, "complementarias": complementarias}


def separar_complementarias(mapa_agrupado, es_caja):
    """
    Dispatcher único: usa el criterio correcto según el tipo de archivo.
    Preferir esta función en vez de llamar directamente a
    separar_complementarias_agrupado / separar_complementarias_agrupado_caja,
    para no repetir el if es_caja en cada caller (ver monitoreo_bot.py).
    """
    if es_caja:
        return separar_complementarias_agrupado_caja(mapa_agrupado)
    return separar_complementarias_agrupado(mapa_agrupado)


# ---------------------------------------------------------------------------
# Generadores de CSV
# ---------------------------------------------------------------------------

def _escribir_csv(lineas, ruta):
    with open(ruta, "w", encoding="utf-8", newline="") as f:
        f.write("\r\n".join(lineas))


def _fila_a_csv(fila, cant_cols):
    cols = []
    for i in range(cant_cols):
        val = fila[i] if i < len(fila) else None
        s = str(val or "").strip()
        if COLS_NUMERICAS and i in COLS_NUMERICAS:
            cols.append(formatear_importe(val))
        else:
            cols.append(s.replace("|", " ").replace("\n", " "))
    return "|".join(cols)


def generar_csv_modificados(modifs, nuevos, ruta):
    cant_cols = CONFIG["COL_FIN"] - CONFIG["COL_INICIO"] + 1
    vistas = set()
    lineas = []

    for c in modifs:
        clave = f"{c.get('id','')}_{c.get('fila','')}"
        if clave not in vistas and c.get("fila"):
            vistas.add(clave)
            lineas.append(_fila_a_csv(c["fila"], cant_cols))

    for c in nuevos:
        if c.get("fila"):
            lineas.append(_fila_a_csv(c["fila"], cant_cols))

    _escribir_csv(lineas, ruta)


def generar_csv_complementarias(complementarias, ruta):
    """Devuelve True si se escribió algo, False si no había complementarias."""
    cant_cols = CONFIG["COL_FIN"] - CONFIG["COL_INICIO"] + 1
    if not complementarias:
        return False

    lineas = []
    for filas in complementarias.values():
        for f in filas:
            lineas.append(_fila_a_csv(f, cant_cols))

    _escribir_csv(lineas, ruta)
    return True


def generar_csv_liquidacion_completa(mapa_agrupado, es_caja, ruta):
    """
    Solo ordinarias (excluye complementarias) en ambos modos.

    NOTA: antes, en modo caja, esta función volcaba TODAS las filas
    (ordinarias + complementarias) porque no había forma de distinguirlas.
    Ahora que separar_complementarias_agrupado_caja() sí las distingue
    (agrupando primero por concepto JUBILADO/PENSIONADO/otro), se alinea
    con el modo normal: la "liquidación completa"/rectificativa queda solo
    con ordinarias, y las complementarias se reportan aparte en su propio
    CSV (generar_csv_complementarias) para no duplicarlas en los dos
    adjuntos.

    NOTA: NO se puede distinguir "ya es una lista de filas" de "es una
    sola fila" con isinstance(filas_o, list) — una fila también es una
    lista (de valores de celda), así que el isinstance da True en ambos
    casos. Hay que usar es_caja, que ya indica cuál de los dos formatos
    devuelve separar_complementarias():
      - modo caja: ordinarias[aid] = lista de filas (list[list])
      - modo normal: ordinarias[aid] = una sola fila (list)
    Con isinstance, en modo normal la fila se tomaba como si ya fuera
    una "lista de filas" y se iteraba directo sobre sus valores (cuil,
    dni, nombre, sueldo...) tratando cada valor como fila completa;
    cuando ese valor era un int, _fila_a_csv explotaba en len(fila).
    """
    cant_cols = CONFIG["COL_FIN"] - CONFIG["COL_INICIO"] + 1
    lineas = []

    resultado = separar_complementarias(mapa_agrupado, es_caja)
    for filas_o in resultado["ordinarias"].values():
        filas = filas_o if es_caja else [filas_o]
        for f in filas:
            lineas.append(_fila_a_csv(f, cant_cols))

    _escribir_csv(lineas, ruta)


# ---------------------------------------------------------------------------
# Generador de XLSX de cambios (con openpyxl)
# ---------------------------------------------------------------------------

# Estilos
_THIN = Side(style="thin", color="D1D5DB")
_MED = Side(style="medium", color="6B7280")
_BRD_N = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BRD_B = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_MED)

_FILL_HEADER = PatternFill("solid", fgColor="074F69")
_FILL_BLUE = PatternFill("solid", fgColor="EFF6FF")
_FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")
_FILL_RED = PatternFill("solid", fgColor="C83C2D")
_FILL_GREEN = PatternFill("solid", fgColor="275317")

_FONT_HDR = Font(name="Calibri", size=11, bold=True, color="8ED973")
_FONT_NORMAL = Font(name="Calibri", size=11)
_FONT_RED = Font(name="Calibri", size=11, color="B91C1C")
_FONT_GREEN = Font(name="Calibri", size=11, bold=True, color="15803D")
_FONT_WHITE = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_FONT_RED_S = Font(name="Calibri", size=11, bold=True, color="C83C2D")
_FONT_GRN_S = Font(name="Calibri", size=11, bold=True, color="275317")
_FONT_BLUE_S = Font(name="Calibri", size=11, bold=True, color="215C98")

_ALIGN_CTR = Alignment(horizontal="center", vertical="center")
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
_NUM_FMT = "#,##0.00"

_ANCHOS = [14, 30, 22, 22, 14, 14, 14, 35] + [16] * 16


def _celda(ws, row, col, valor, font=None, fill=None, border=None, align=None, num_fmt=None):
    c = ws.cell(row=row, column=col, value=valor)
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if border:
        c.border = border
    if align:
        c.alignment = align
    if num_fmt:
        c.number_format = num_fmt
    return c


def generar_xlsx_cambios(modifs, elims, nuevos, periodo, reparticion, ruta_salida):
    wb = openpyxl.Workbook()
    ws = wb.active
    nombre_hoja = re.sub(r'[\\/:*?\[\]]', '-', str(periodo)).replace("°", "")[:31].strip()
    ws.title = nombre_hoja

    for i, ancho in enumerate(_ANCHOS, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = ancho

    row = 1

    # ── REGISTROS MODIFICADOS ──
    if modifs:
        _celda(ws, row, 1, "REGISTROS MODIFICADOS", font=_FONT_BLUE_S)
        row += 1

        hdrs = ["DNI", "Nombre y Apellido", "Campo modificado", "Valor anterior", "Valor nuevo"]
        for c_idx, h in enumerate(hdrs, 1):
            _celda(ws, row, c_idx, h, font=_FONT_HDR, fill=_FILL_HEADER, border=_BRD_N, align=_ALIGN_CTR)
        row += 1

        # Agrupar por DNI
        grupos = {}
        orden_dni = []
        for c in modifs:
            d = c["dni"]
            if d not in grupos:
                grupos[d] = {"nombre": c["nombre"], "cambios": []}
                orden_dni.append(d)
            grupos[d]["cambios"].append(c)

        for gi, dni in enumerate(orden_dni):
            g = grupos[dni]
            fill = _FILL_BLUE if gi % 2 == 0 else _FILL_WHITE
            cc = g["cambios"]
            for i, c in enumerate(cc):
                es_ult = i == len(cc) - 1
                brd = _BRD_B if es_ult else _BRD_N
                campo = c["columna"]
                idx_c = NOMBRES_COLUMNAS.index(campo) if campo in NOMBRES_COLUMNAS else -1
                campo_label = f"{idx_c+1}-{campo}" if idx_c >= 0 else campo

                _celda(ws, row, 1, dni if i == 0 else "", font=_FONT_NORMAL, fill=fill, border=brd, align=_ALIGN_LEFT)
                _celda(ws, row, 2, g["nombre"] if i == 0 else "", font=_FONT_NORMAL, fill=fill, border=brd, align=_ALIGN_LEFT)
                _celda(ws, row, 3, campo_label, font=_FONT_NORMAL, fill=fill, border=brd, align=_ALIGN_LEFT)

                if c["es_no_numerico"]:
                    _celda(ws, row, 4, str(c["anterior"]), font=_FONT_RED, fill=fill, border=brd, align=_ALIGN_LEFT)
                    _celda(ws, row, 5, str(c["actual"]), font=_FONT_GREEN, fill=fill, border=brd, align=_ALIGN_LEFT)
                else:
                    _celda(ws, row, 4, parse_numero(c["anterior"]), font=_FONT_RED, fill=fill, border=brd, align=_ALIGN_LEFT, num_fmt=_NUM_FMT)
                    _celda(ws, row, 5, parse_numero(c["actual"]), font=_FONT_GREEN, fill=fill, border=brd, align=_ALIGN_LEFT, num_fmt=_NUM_FMT)
                row += 1
        row += 1

    # ── Helper para secciones de filas completas ──
    def _seccion(titulo, lista, font_titulo, fill_fila, font_fila):
        nonlocal row
        _celda(ws, row, 1, titulo, font=font_titulo)
        row += 1
        for ci, enc in enumerate(ENCABEZADOS_EXCEL, 1):
            _celda(ws, row, ci, enc, font=_FONT_HDR, fill=_FILL_HEADER, border=_BRD_N, align=_ALIGN_CTR)
        row += 1
        for i, c in enumerate(lista):
            brd = _BRD_B if i == len(lista) - 1 else _BRD_N
            fila = c.get("fila", [])
            cant = CONFIG["COL_FIN"] - CONFIG["COL_INICIO"] + 1
            for ci in range(cant):
                val = fila[ci] if ci < len(fila) else None
                if ci in COLS_NUMERICAS:
                    _celda(ws, row, ci+1, parse_numero(val), font=font_fila, fill=fill_fila, border=brd, align=_ALIGN_CTR, num_fmt=_NUM_FMT)
                else:
                    _celda(ws, row, ci+1, str(val or ""), font=font_fila, fill=fill_fila, border=brd, align=_ALIGN_LEFT)
            row += 1
        row += 1

    # Orden de secciones: Modificados (ya generado arriba) -> Nuevos -> Eliminados
    if nuevos:
        _seccion("REGISTROS NUEVOS", nuevos, _FONT_GRN_S, _FILL_GREEN, _FONT_WHITE)

    if elims:
        _seccion("REGISTROS ELIMINADOS", elims, _FONT_RED_S, _FILL_RED, _FONT_WHITE)

    wb.save(ruta_salida)


# ---------------------------------------------------------------------------
# Helpers de nombres / periodos
# ---------------------------------------------------------------------------

_MESES = {
    "01": "Enero", "02": "Febrero", "03": "Marzo", "04": "Abril",
    "05": "Mayo", "06": "Junio", "07": "Julio", "08": "Agosto",
    "09": "Septiembre", "10": "Octubre", "11": "Noviembre", "12": "Diciembre",
}


def hoja_a_periodo(hoja, anio):
    hl = hoja.strip().lower().replace("º", "°")
    if hl in ("1° sac", "1°sac"):
        return f"1° SAC/{anio}"
    if hl in ("2° sac", "2°sac"):
        return f"2° SAC/{anio}"
    h = hoja.strip()
    if h in _MESES:
        return f"{_MESES[h]}/{anio}"
    return f"{h}/{anio}"


def extraer_reparticion(nombre_archivo):
    sin_ext = nombre_archivo.replace(".xlsx", "").replace(".XLSX", "")
    partes = sin_ext.split("-")
    if len(partes) >= 3:
        ultimo = partes[-1].strip()
        fin = len(partes) - 1 if re.match(r"^\d{4}$", ultimo) else len(partes)
        return "-".join(partes[1:fin]).strip()
    return sin_ext


def extraer_anio_desde_nombre(nombre):
    m = re.search(r"(19\d{2}|20\d{2})", nombre)
    return int(m.group(1)) if m else datetime.now().year


def normalizar_nombre(r):
    return "".join(w.capitalize() for w in re.split(r"[\s\-]+", r.lower()))


def normalizar_periodo(p):
    return re.sub(r"[^a-zA-Z0-9]", "", p.replace("°", "").replace("º", ""))


def formatear_lista_periodos(periodos):
    """
    Junta una lista de períodos en un texto, en MAYÚSCULAS, con el mismo
    criterio que tenía el Apps Script original:
      - 1 o 2 períodos: unidos con " y "        -> "MAYO/2026 y JUNIO/2026"
      - 3 o más: coma entre todos menos el último, " y " antes del último
                                                  -> "MAYO/2026, JUNIO/2026 y 1° SAC/2026"
    """
    periodos_mayus = [p.upper() for p in periodos]
    if len(periodos_mayus) <= 2:
        return " y ".join(periodos_mayus)
    return ", ".join(periodos_mayus[:-1]) + " y " + periodos_mayus[-1]


def construir_asunto_monitoreo(reparticion, periodos):
    """
    Arma el asunto del mail de cambios.
    Ej: "🔄🔵🟢 RECTIFICATIVA | OSER - MINISTERIO DE SALUD PERIODOS: MAYO/2026, JUNIO/2026 y 1° SAC/2026"
    """
    label_periodo = "PERIODO" if len(periodos) == 1 else "PERIODOS"
    periodos_str = formatear_lista_periodos(periodos)
    return f"🔄🔵🟢 RECTIFICATIVA | OSER - {reparticion.upper()} {label_periodo}: {periodos_str}"