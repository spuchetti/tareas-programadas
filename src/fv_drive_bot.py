"""
Bot de Fondo Voluntario - Busca archivos con diferencias y genera CSV con delimitador |
"""
"""
======= EJECUCIÓN MANUAL =========
Modificar/Revisar:
MES_ACTUAL
obtener_anio() -> anio_actual

"""
import openpyxl
import sys
import os
import time
import concurrent.futures
from datetime import datetime
from io import BytesIO

# Para que encuentre los módulos utils
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# imports
from utils.common_utils import (
    registrar_inicio, registrar_resumen, 
    nombre_mes, obtener_mes_anterior, obtener_anio, crear_directorio_salida
)
from utils.drive_utils import inicializar_drive, obtener_archivos, descargar_archivo
from utils.gmail_utils import enviar_email_html_con_adjuntos, generar_html_resumen_fv


# Configuración
MES_ACTUAL = obtener_mes_anterior()
ANIO_ACTUAL = obtener_anio(MES_ACTUAL)

COLUMNAS_REVISION = [52, 53, 54]  # AZ, BA, BB
FILA_INICIO = 4
MAXIMO_HILOS = 5

# Columnas a extraer para el CSV
COLUMNAS_EXTRACCION = {
    'dni': 2,           # Columna B
    'reparticion': 8,   # Columna H
    'cuota': 51,        # Columna AY
    'omision_total': 52, # Columna AZ
    'aporte_inferior': 53, # Columna BA
    'supera_cuota': 54   # Columna BB
}


def valor_a_float(valor):
    """
    Convierte un valor de celda a float, manejando casos especiales.
    Si el valor es "-", retorna 0.00
    """
    if valor is None:
        return 0.00
    
    if isinstance(valor, (int, float)):
        return float(valor)
    
    if isinstance(valor, str):
        valor_str = valor.strip()
        if valor_str == "-" or valor_str == "":
            return 0.00
        try:
            # Intentar convertir a float (reemplaza coma por punto si es necesario)
            valor_limpio = valor_str.replace(',', '.')
            return float(valor_limpio)
        except ValueError:
            return 0.00
    
    return 0.00


def buscar_en_hoja(fh, nombre, hoja, todas_las_filas):
    """
    Busca en la hoja las filas donde alguna de las columnas AZ, BA, BB tenga valor != 0
    (tanto positivos como negativos)
    Acumula los datos extraídos en la lista 'todas_las_filas' que se pasa por referencia
    
    Args:
        fh: File handle del archivo Excel
        nombre: Nombre del archivo
        hoja: Nombre de la hoja a revisar
        todas_las_filas: Lista donde se acumulan todas las filas encontradas (se modifica in-place)
    
    Returns:
        bool: True si se encontró al menos una fila que cumple la condición
    """
    try:
        fh.seek(0)
        wb = openpyxl.load_workbook(fh, data_only=True, read_only=True)

        if hoja not in wb.sheetnames:
            print(f"   ⚠ Hoja '{hoja}' no encontrada en {nombre}")
            wb.close()
            return False

        ws = wb[hoja]
        filas_encontradas_en_archivo = 0

        # DEBUG: Muestra primeras filas para verificar
        print(f"   🔍 Buscando en hoja '{hoja}' desde fila {FILA_INICIO}")

        for row_idx, row in enumerate(ws.iter_rows(min_row=FILA_INICIO, values_only=True), start=FILA_INICIO):
            if row is None:
                break

            # Verifica si llegamos al final (fila vacía o con "-" en columna A)
            if len(row) > 0:
                primera_celda = row[0]
                if primera_celda is None or str(primera_celda).strip() == "" or str(primera_celda).strip() == "-":
                    break

            # Verifica si la fila tiene suficientes columnas
            if len(row) < max(COLUMNAS_REVISION):
                continue

            # Verifica valores en las columnas de revisión
            hay_valor_diferente_de_cero = False
            valores_revision = []
            
            for col_idx in COLUMNAS_REVISION:
                if len(row) >= col_idx:
                    valor = row[col_idx - 1]
                    valor_float = valor_a_float(valor)
                    valores_revision.append(valor_float)
                    if valor_float != 0:
                        hay_valor_diferente_de_cero = True

            # Solo procesar si hay al menos un valor diferente de cero
            if hay_valor_diferente_de_cero:
                # Extrae los datos requeridos
                dni = ""
                reparticion = ""
                cuota = 0.00
                omision_total = valores_revision[0] if len(valores_revision) > 0 else 0.00
                aporte_inferior = valores_revision[1] if len(valores_revision) > 1 else 0.00
                supera_cuota = valores_revision[2] if len(valores_revision) > 2 else 0.00
                
                # Extrae DNI (columna B)
                if len(row) >= COLUMNAS_EXTRACCION['dni']:
                    dni_val = row[COLUMNAS_EXTRACCION['dni'] - 1]
                    if dni_val is not None:
                        dni = str(dni_val).strip()
                        # Si es número, elimina decimales
                        if dni.replace('.', '').isdigit():
                            dni = dni.split('.')[0]
                
                # Extrae REPARTICIÓN (columna H)
                if len(row) >= COLUMNAS_EXTRACCION['reparticion']:
                    rep_val = row[COLUMNAS_EXTRACCION['reparticion'] - 1]
                    if rep_val is not None:
                        reparticion = str(rep_val).strip()
                
                # Extrae CUOTA (columna AY)
                if len(row) >= COLUMNAS_EXTRACCION['cuota']:
                    cuota = valor_a_float(row[COLUMNAS_EXTRACCION['cuota'] - 1])
                
                # Solo agrega si hay DNI válido
                if dni and dni != "" and dni != "None":
                    todas_las_filas.append((dni, reparticion, cuota, omision_total, aporte_inferior, supera_cuota))
                    filas_encontradas_en_archivo += 1
                    
                    # DEBUG: Muestra primeras filas encontradas
                    if filas_encontradas_en_archivo <= 3:
                        print(f"   ✅ Fila {row_idx}: DNI={dni}, Repartición={reparticion[:30]}..., Cuota={cuota:.2f}")

        wb.close()
        
        if filas_encontradas_en_archivo > 0:
            print(f"   📊 Archivo {nombre}: {filas_encontradas_en_archivo} fila(s) con valores ≠ 0")
            return True
        else:
            return False

    except Exception as e:
        print(f"❌ Error leyendo {nombre}: {e}")
        return False


def generar_archivo_csv_unico(filas_encontradas, periodo, anio):
    """
    Genera un archivo CSV con todos los casos encontrados.
    Formato: DNI|REPARTICION|CUOTA|OMISION_TOTAL|APORTE_INFERIOR|SUPERA_CUOTA (incluido el encabezado)
    Delimitador: pipe (|)
    
    Args:
        filas_encontradas: Lista de tuplas (dni, reparticion, cuota, omision_total, aporte_inferior, supera_cuota)
        periodo: Período que se está procesando
        anio: Año actual
    
    Returns:
        Ruta del archivo CSV generado o None si no hay filas
    """
    try:
        if not filas_encontradas:
            return None
        
        # Crea nombre del archivo CSV
        nombre_mes_legible = nombre_mes(periodo)
        nombre_csv = f"Casos_FondoVoluntario_{nombre_mes_legible}{anio}.csv"
        
        carpeta = crear_directorio_salida()
        ruta = os.path.join(carpeta, nombre_csv)
        
        with open(ruta, 'w', encoding='utf-8') as f:
            # encabezados
            encabezados = "DNI|REPARTICION|CUOTA|OMISION_TOTAL|APORTE_INFERIOR|SUPERA_CUOTA\n"
            f.write(encabezados)
            
            # Escribe los datos
            for fila in filas_encontradas:
                dni, reparticion, cuota, omision, aporte, supera = fila
                
                # Limpia el campo repartición para que no contenga pipes
                reparticion_limpia = reparticion.replace('|', '-')
                
                # Formato CSV con pipe como delimitador
                linea = f"{dni}|{reparticion_limpia}|{cuota:.2f}|{omision:.2f}|{aporte:.2f}|{supera:.2f}\n"
                f.write(linea)
        
        print(f"\n📄 Archivo CSV generado: {nombre_csv}")
        print(f"   Total de casos en el archivo: {len(filas_encontradas)}")
        print(f"   Formato: DNI|REPARTICION|CUOTA|OMISION|APORTE_INFERIOR|SUPERA_CUOTA")
        print(f"   Delimitador: pipe (|)")
        print(f"   Ruta: {ruta}")
        
        # Mostrar primeras líneas como ejemplo
        with open(ruta, 'r', encoding='utf-8') as f:
            primeras_lineas = f.readlines()[:3]
            print(f"\n   📊 Primeras líneas del CSV:")
            for linea in primeras_lineas:
                print(f"     {linea.strip()}")
        
        return ruta
        
    except Exception as e:
        print(f"❌ Error generando archivo CSV único: {e}")
        return None


def procesar_archivo(archivo, hoja_periodo, todas_las_filas):
    """
    Procesa un archivo individual - busca filas con valores != 0 en AZ, BA, BB
    Acumula los datos en todas_las_filas
    
    Args:
        archivo: Información del archivo de Drive
        hoja_periodo: Período a procesar
        todas_las_filas: Lista acumuladora de todas las filas encontradas
    
    Returns:
        tuple: (nombre_archivo, tiene_casos)
    """
    try:
        servicio_drive = inicializar_drive()
        if not servicio_drive:
            return archivo["name"], False

        fh = descargar_archivo(servicio_drive, archivo)
        if not fh:
            return archivo["name"], False

        # Busca filas que cumplan la condición y acumular en todas_las_filas
        tiene_casos = buscar_en_hoja(fh, archivo["name"], hoja_periodo, todas_las_filas)

        return archivo["name"], tiene_casos
        
    except Exception as e:
        print(f"❌ Error procesando {archivo['name']}: {e}")
        return archivo["name"], False


def ejecutar_principal():
    """Función principal del bot"""
    inicio = time.time()
    ahora = registrar_inicio("BOT FONDO VOLUNTARIO")

    # 1. Inicializa el Drive
    drive = inicializar_drive()
    if not drive:
        return

    # 2. Obtiene archivos
    archivos = obtener_archivos(drive)
    if not archivos:
        print("❌ No se encontraron archivos.")
        return

    # 3. Determina período
    periodo = MES_ACTUAL
    anio_actual = ANIO_ACTUAL
    periodo_legible = f"{nombre_mes(periodo)}/{anio_actual}"
    periodo_legible_upper = periodo_legible.upper()

    print(f"📄 Hoja a controlar: {periodo} ({periodo_legible})")
    print(f"🔍 Criterio: Buscar valores != 0 en columnas AZ, BA, BB (tanto positivos como negativos)")
    print(f"📋 Se generará un ÚNICO archivo CSV con TODOS los casos")
    print(f"📋 Formato CSV: DNI|REPARTICION|CUOTA|OMISION_TOTAL|APORTE_INFERIOR|SUPERA_CUOTA")
    print(f"📋 Delimitador: pipe (|)\n")

    # 4. Lista acumuladora de todas las filas encontradas en todos los archivos
    todas_las_filas = []
    
    # 5. Procesa archivos en paralelo
    archivos_con_casos = []  # Archivos que tienen al menos una fila que cumple
    archivos_sin_casos = []  # Archivos que no tienen ninguna fila

    with concurrent.futures.ThreadPoolExecutor(max_workers=MAXIMO_HILOS) as pool:
        # Pasa todas_las_filas como argumento a cada tarea
        tareas = [
            pool.submit(procesar_archivo, a, periodo, todas_las_filas)
            for a in archivos
        ]

        for future in concurrent.futures.as_completed(tareas):
            nombre, tiene_casos = future.result()

            if tiene_casos:
                print(f"   ✔ {nombre} → TIENE CASOS")
                archivos_con_casos.append(nombre)
            else:
                print(f"   ✖ {nombre}")
                archivos_sin_casos.append(nombre)

    # 6. Genera archivo CSV con todos los casos encontrados
    ruta_csv_unico = generar_archivo_csv_unico(todas_las_filas, periodo, anio_actual)

    # 7. Prepara lista de adjuntos (solo CSV)
    adjuntos = []
    if ruta_csv_unico:
        adjuntos.append(ruta_csv_unico)

    # 8. Envia email (si hay archivos)
    if adjuntos:
        html = generar_html_resumen_fv(
            periodo_legible,
            len(archivos),
            len(archivos_con_casos),
            len(todas_las_filas),
            archivos_con_casos,
            ahora.strftime("%d-%m-%Y %H:%M:%S")
        )

        asunto = f"🟢🔵 OSER - CONTROL AUTOMÁTICO FONDO VOLUNTARIO | PERIODO: {periodo_legible_upper}"
        enviar_email_html_con_adjuntos(asunto, html, adjuntos, "SMTP_TO_FV")
    else:
        print("⚠ No se generaron archivos para adjuntar")

    #  Muestra resumen detallado
    print("\n" + "=" * 70)
    print("📊 RESUMEN FINAL")
    print("=" * 70)
    print(f"📁 Archivos procesados: {len(archivos)}")
    print(f"✅ Reparticiones detectadas: {len(archivos_con_casos)}")
    print(f"❌ Archivos sin casos: {len(archivos_sin_casos)}")
    print(f"📋 Total de agentes detectados: {len(todas_las_filas)}")
    print(f"📄 Archivo CSV único generado: {'SÍ' if ruta_csv_unico else 'NO'}")
    print(f"📎 Total adjuntos en email: {len(adjuntos)}")
    
    # 9. Muestra los primeros 5 casos como ejemplo
    if todas_las_filas:
        print("\n📋 Ejemplos de casos encontrados (primeros 5):")
        for i, fila in enumerate(todas_las_filas[:5], 1):
            dni, reparticion, cuota, omision, aporte, supera = fila
            reparticion_corta = reparticion[:30] + "..." if len(reparticion) > 30 else reparticion
            print(f"   {i}. DNI: {dni} | Rep: {reparticion_corta} | Cuota: {cuota:.2f}")
    
    print("=" * 70)
    
    registrar_resumen(inicio, len(archivos), len(archivos_con_casos))


if __name__ == "__main__":
    ejecutar_principal()