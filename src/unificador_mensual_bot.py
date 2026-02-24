"""
Unificador Mensual - Actualiza CSV consolidado en Drive agregando datos del mes actual
"""
"""
======= EJECUCIÓN MANUAL =========
Modificar/Revisar:
MES_ACTUAL
obtener_anio() -> anio_actual

enviar_email_html_adjuntos():
"SMTP_TO_UNIFICADOR"(normal)
"SMTP_TO_FV"(prueba unitaria)


"""

import openpyxl
import sys
import os
import io
import csv
import time
import re
import traceback
from datetime import datetime

# Agregar esta línea para que Python encuentre los módulos utils
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Luego el resto de imports
from utils.common_utils import (
    registrar_inicio, registrar_resumen, 
    nombre_mes, obtener_mes_anterior, obtener_anio
)
from utils.drive_utils import (
    inicializar_drive, obtener_archivos, descargar_archivo,
    guardar_csv_localmente
)
from utils.excel_utils import eliminar_tildes_latin, normalizar_texto
from utils.gmail_utils import (
    enviar_email_html_con_adjuntos, 
    generar_html_resumen_unificador
)

# Configuración
MES_ACTUAL = obtener_mes_anterior()  # Mes que estamos procesando


def obtener_nombre_csv():
    """Obtiene el nombre del archivo CSV basado en el mes y año actual"""
    mes_actual_formateado = MES_ACTUAL
    anio_actual = obtener_anio(mes_actual_formateado)
    nombre_mes_actual = nombre_mes(mes_actual_formateado)
    
    # Formato: Unificado_MesAño.csv (ej: Unificado_Noviembre2025.csv)
    nombre_csv = f"Unificado_{nombre_mes_actual}{anio_actual}.csv"
    return nombre_csv

def extraer_codigo_desde_nombre(nombre_archivo):
    """
    Extrae el código del nombre del archivo Excel.
    Formato esperado: codigo-nombre de reparticion-año.xlsx
    
    Args:
        nombre_archivo: Nombre completo del archivo (ej: "1234-Municipalidad de Capital-2025.xlsx")
    
    Returns:
        str: Código extraído (ej: "1234")
    """
    try:
        # Obtener nombre sin extensión
        nombre_sin_extension = os.path.splitext(nombre_archivo)[0]
        
        # El código es lo que está antes del primer guión
        partes = nombre_sin_extension.split('-', 1)
        if partes:
            codigo = partes[0].strip()
            # Validar que el código no esté vacío
            if codigo and codigo != "":
                print(f"   🔑 Código extraído: '{codigo}' de '{nombre_archivo}'")
                return codigo
            else:
                print(f"   ⚠ Código vacío en '{nombre_archivo}', usando 'SIN_CODIGO'")
                return "SIN_CODIGO"
        else:
            print(f"   ⚠ No se pudo extraer código de '{nombre_archivo}', usando 'SIN_CODIGO'")
            return "SIN_CODIGO"
            
    except Exception as e:
        print(f"   ❌ Error extrayendo código de '{nombre_archivo}': {e}")
        return "SIN_CODIGO"

def extraer_datos_excel(fh, nombre_archivo, hoja_mes):
    """
    Extrae datos de un archivo Excel desde fila 4, columnas A to X
    en la hoja anterior al mes actual.
    Se detiene cuando encuentra "-" o celda vacía en columna A.
    AHORA TAMBIÉN EXTRAE EL CÓDIGO DEL NOMBRE DEL ARCHIVO PARA LA COLUMNA 25.
    
    NOTA: Para archivos de "Caja" empieza desde fila 5.
    
    COLUMNAS ESPECIALES:
    - D (4): Nombre y apellido - ELIMINAR TILDES (SÍ eliminar)
    - F (6): Situación de revista - ELIMINAR TILDES (SÍ eliminar)
    - H (8): Repartición - ELIMINAR TILDES (SÍ eliminar)
    - I-X (9-24): Números con 2 decimales
    - Otras columnas de texto: mantener tildes
    """
    try:
        # Extraer código del nombre del archivo
        codigo_archivo = extraer_codigo_desde_nombre(nombre_archivo)
        
        # Cargar libro
        fh.seek(0)
        wb = openpyxl.load_workbook(fh, data_only=True, read_only=True)
        
        # Mostrar todas las hojas disponibles para debug
        print(f"   📋 Hojas disponibles en {nombre_archivo}: {wb.sheetnames}")

        # Verificar si existe la hoja del mes
        if hoja_mes not in wb.sheetnames:
            print(f"⚠ Hoja '{hoja_mes}' no encontrada en {nombre_archivo}")
            # Mostrar hojas similares
            hojas_similares = [s for s in wb.sheetnames if hoja_mes.lower() in s.lower()]
            if hojas_similares:
                print(f"   ℹ️ Hojas similares encontradas: {hojas_similares}")

            wb.close()
            return []
        
        ws = wb[hoja_mes]
        datos_extraidos = []
        
        # DETERMINAR FILA DE INICIO
        # Si es archivo de "Caja", empezar desde fila 5, sino desde fila 4
        es_caja = "caja" in nombre_archivo.lower()
        if es_caja:
            fila_inicio = 5
            print(f"   ⚙ Archivo 'Caja' detectado. Iniciando desde fila {fila_inicio}")
        else:
            fila_inicio = 4
        
        # Procesar desde la fila de inicio
        for row_idx, row in enumerate(ws.iter_rows(min_row=fila_inicio, max_col=24, values_only=True), start=fila_inicio):
            
            # VERIFICAR CONDICIÓN DE PARADA: "-" o celda vacía en columna A
            primera_celda = row[0] if len(row) > 0 else None
            
            # Convertir a string para la verificación
            if primera_celda is None:
                primera_celda_str = ""
            elif isinstance(primera_celda, datetime):
                primera_celda_str = primera_celda.strftime("%Y-%m-%d")
            else:
                primera_celda_str = str(primera_celda).strip()
            
            # CONDICIÓN DE PARADA GENERAL
            if primera_celda_str == "-" or primera_celda_str == "":
                print(f"   ⏹ Marcador de fin encontrado en fila {row_idx} ('{primera_celda_str}'). Fin de extracción.")
                wb.close()
                return datos_extraidos
            
            # También verificar si todas las celdas de la fila están vacías
            if all(cell is None or cell == '' or str(cell).strip() == '' for cell in row):
                print(f"   ⏹ Fila {row_idx} completamente vacía. Fin de extracción.")
                wb.close()
                return datos_extraidos
            
            # LIMPIAR Y FORMATAR CADA CELDA
            fila_limpia = []
            for col_idx, cell in enumerate(row, start=1):
                if cell is None:
                    fila_limpia.append("")
                elif isinstance(cell, datetime):
                    fila_limpia.append(cell.strftime("%Y-%m-%d"))
                else:
                    # Columnas A-H (1-8): formatear como texto
                    if col_idx <= 8:  
                        # Para columnas de texto
                        if isinstance(cell, (int, float)):
                            # Si es número (CUIL/DNI), eliminar decimales innecesarios
                            if isinstance(cell, float) and cell.is_integer():
                                fila_limpia.append(str(int(cell)))
                            else:
                                cell_str = str(cell)
                                # Eliminar .0 final si existe
                                if cell_str.endswith('.0'):
                                    cell_str = cell_str[:-2]
                                fila_limpia.append(cell_str)
                        else:
                            # Para texto
                            cell_str = str(cell) if cell is not None else ""
                            
                            # IMPORTANTE: Columnas D (4), F (6) y H (8) - ELIMINAR TILDES
                            if col_idx == 4 or col_idx == 6 or col_idx == 8:
                                # COLUMNA D (Nombre y Apellido), F (Situación revista) y H (Repartición) - ELIMINAR TILDES
                                fila_limpia.append(normalizar_texto(cell_str, eliminar_tildes_param=True))
                            else:
                                # Otras columnas de texto - mantener tildes
                                fila_limpia.append(normalizar_texto(cell_str, eliminar_tildes_param=False))
                    
                    else:  # Columnas I-X (9-24) - números SIEMPRE con 2 decimales y punto
                        if cell == "" or cell is None:
                            # Para celdas vacías, poner "0.00"
                            fila_limpia.append("0.00")
                        elif isinstance(cell, (int, float)):
                            # Para números enteros o decimales, formatear siempre con 2 decimales
                            formatted = f"{float(cell):.2f}"
                            fila_limpia.append(formatted)
                        else:
                            # Si es texto, intentar convertir a número
                            cell_str = str(cell).strip()
                            
                            if cell_str == "" or cell_str.lower() == "nan":
                                fila_limpia.append("0.00")
                            elif cell_str == "0":
                                fila_limpia.append("0.00")
                            else:
                                try:
                                    # Normalizar: reemplazar comas por puntos
                                    cell_normalized = cell_str.replace(',', '.')
                                    
                                    # Manejar múltiples puntos
                                    if cell_normalized.count('.') > 1:
                                        parts = cell_normalized.split('.')
                                        integer_part = ''.join(parts[:-1])
                                        decimal_part = parts[-1]
                                        cell_normalized = f"{integer_part}.{decimal_part}"
                                    
                                    num = float(cell_normalized)
                                    formatted = f"{num:.2f}"
                                    fila_limpia.append(formatted)
                                    
                                except (ValueError, AttributeError):
                                    fila_limpia.append("0.00")
            
            # Solo agregar si la fila tiene algún contenido
            if any(cell != "" for cell in fila_limpia):
                # AGREGAR CÓDIGO COMO COLUMNA 25
                fila_con_codigo = fila_limpia + [codigo_archivo]
                datos_extraidos.append(fila_con_codigo)
        
        wb.close()
        
        print(f"📊 Extraídos {len(datos_extraidos)} filas de {nombre_archivo} (hoja {hoja_mes}, desde fila {fila_inicio})")
        print(f"   🔑 Código aplicado a todas las filas: '{codigo_archivo}'")
        
        # Mostrar ejemplos detallados para debug
        if datos_extraidos and len(datos_extraidos) > 0:
            primera_fila = datos_extraidos[0]
            print(f"   🔍 Verificación de formato (primer registro):")
            
            if len(primera_fila) > 3:
                # Columna D (Nombre)
                nombre_original = primera_fila[3]
                nombre_sin_tildes = eliminar_tildes_latin(nombre_original) if nombre_original else ""
                print(f"     Col D (Nombre original): '{nombre_original}'")
                print(f"     Col D (Nombre sin tildes): '{nombre_sin_tildes}'")
            
            if len(primera_fila) > 3:
            # Columna F (Situación de revista)
                situacion_original = primera_fila[6]
                situacion_sin_tildes = eliminar_tildes_latin(situacion_original) if situacion_original else ""
                print(f"     Col F (Nombre original): '{situacion_original}'")
                print(f"     Col F (Nombre sin tildes): '{situacion_sin_tildes}'")

            if len(primera_fila) > 7:
                # Columna H (Repartición)
                reparticion_original = primera_fila[7]
                reparticion_sin_tildes = eliminar_tildes_latin(reparticion_original) if reparticion_original else ""
                print(f"     Col H (Repartición original): '{reparticion_original}'")
                print(f"     Col H (Repartición sin tildes): '{reparticion_sin_tildes}'")
            
            # Mostrar código agregado
            if len(primera_fila) > 24:
                print(f"     Col 25 (Código): '{primera_fila[24]}'")
            
            # Mostrar ejemplo de conversión
            if len(primera_fila) > 3 and "á" in primera_fila[3] or "é" in primera_fila[3] or "í" in primera_fila[3] or "ó" in primera_fila[3] or "ú" in primera_fila[3]:
                print(f"     ✅ Ejemplo de conversión de tildes aplicado correctamente")
        
        return datos_extraidos
        
    except Exception as e:
        print(f"❌ Error extrayendo datos de {nombre_archivo}: {e}")
        import traceback
        traceback.print_exc()
        return []

def determinar_tipo_reparticion(nombre_archivo):
    """
    Determina el tipo de entidad basado en el nombre del archivo.
    Versión MEJORADA con matching más preciso usando expresiones regulares.
    SI NO ENCUENTRA COINCIDENCIA, DEVUELVE 'Otros'
    """
    nombre_lower = nombre_archivo.lower()
    nombre_sin_ext = os.path.splitext(nombre_archivo)[0].lower()
    nombre_sin_ext = re.sub(r'[-_\s]+', ' ', nombre_sin_ext).strip()
    
    # Patrones específicos para cada tipo
    municipio_patterns = [
        r'municipio',
        r'municipalidad',
        r'intendencia',
        r'municipal',
        r'^mun\.',
        r'municipio de'
    ]
    
    comuna_patterns = [
        r'comuna',
        r'^com\.',
        r'comuna de'
    ]
    
    entes_patterns = [
        r'ente',
        r'ente.*descentralizado',
        r'ente.*autarq',
        r'autarquico',
        r'autárquico',
        r'instituto.*autarq',
        r'organismo.*descentralizado',
        r'entidad.*autonoma',
        r'entidad.*autónoma',
        r'autarq'
    ]
    
    cajas_patterns = [
        r'caja',
        r'caja.*municipal',
        r'caja.*provincial',
        r'banco.*municipal',
        r'caja de jubilaciones',
        r'caja de previsión',
        r'caja de prevision'
    ]
    
    escuela_patterns = [
        r'idessa',
        r'escuela',
        r'instituto.*educacion',
        r'instituto.*educación',
        r'colegio',
        r'universidad',
        r'facultad'
    ]
    
    # Verificar patrones en orden de prioridad
    for pattern in cajas_patterns:
        if re.search(pattern, nombre_lower) or re.search(pattern, nombre_sin_ext):
            print(f"[CLASIFICACIÓN MEJORADA] Archivo: '{nombre_archivo}' → Cajas Municipales (patrón: {pattern})")
            return 'Cajas Municipales'
    
    for pattern in escuela_patterns:
        if re.search(pattern, nombre_lower) or re.search(pattern, nombre_sin_ext):
            print(f"[CLASIFICACIÓN MEJORADA] Archivo: '{nombre_archivo}' → Escuela (patrón: {pattern})")
            return 'Escuela'
    
    for pattern in entes_patterns:
        if re.search(pattern, nombre_lower) or re.search(pattern, nombre_sin_ext):
            print(f"[CLASIFICACIÓN MEJORADA] Archivo: '{nombre_archivo}' → Entes Descentralizados (patrón: {pattern})")
            return 'Entes Descentralizados'
    
    for pattern in comuna_patterns:
        if re.search(pattern, nombre_lower) or re.search(pattern, nombre_sin_ext):
            print(f"[CLASIFICACIÓN MEJORADA] Archivo: '{nombre_archivo}' → Comunas (patrón: {pattern})")
            return 'Comunas'
    
    for pattern in municipio_patterns:
        if re.search(pattern, nombre_lower) or re.search(pattern, nombre_sin_ext):
            print(f"[CLASIFICACIÓN MEJORADA] Archivo: '{nombre_archivo}' → Municipios (patrón: {pattern})")
            return 'Municipios'
    
    # SI NO ENCUENTRA NINGÚN PATRÓN, DEVOLVER 'Otros'
    print(f"[CLASIFICACIÓN MEJORADA] Archivo: '{nombre_archivo}' → Otros (no se encontró patrón)")
    return 'Otros'

def verificar_consistencia_sumatorias_detallada(periodo, sumatorias_por_tipo, ruta_csv):
    """
    Verifica en detalle la consistencia de las sumatorias para un período específico.
    """
    from utils.gmail_utils import calcular_sumatorias_csv
    
    print(f"\n{'='*70}")
    print(f"🔍 VERIFICACIÓN DETALLADA DE CONSISTENCIA - PERÍODO {periodo}")
    print(f"{'='*70}")
    
    # Calcular sumatorias del CSV
    print("📊 Calculando sumatorias del CSV...")
    sumatorias_csv = calcular_sumatorias_csv(ruta_csv)
    
    # Calcular sumatorias acumuladas de los tipos
    print("📈 Calculando sumatorias acumuladas por tipo de entidad...")
    sumatorias_acumuladas = {
        'creditos_asistenciales': 0.0,
        'fondo_voluntario': 0.0,
        'personal': 0.0,
        'adherente': 0.0,
        'patronal': 0.0,
        'total': 0.0
    }
    
    # Contador de tipos con datos
    tipos_con_datos = 0
    
    # Acumular por tipo (excluyendo 'Otro')
    for tipo_entidad, sumatorias in sumatorias_por_tipo.items():
        if tipo_entidad != 'Otro' and sumatorias['total'] > 0:
            tipos_con_datos += 1
            print(f"  ➕ Acumulando {tipo_entidad}: ${sumatorias['total']:,.2f}")
            
            for concepto in sumatorias_acumuladas:
                if concepto in sumatorias:
                    sumatorias_acumuladas[concepto] += sumatorias[concepto]
    
    # Calcular total acumulado
    sumatorias_acumuladas['total'] = (
        sumatorias_acumuladas['creditos_asistenciales'] +
        sumatorias_acumuladas['fondo_voluntario'] +
        sumatorias_acumuladas['personal'] +
        sumatorias_acumuladas['adherente'] +
        sumatorias_acumuladas['patronal']
    )
    
    print(f"\n📊 RESULTADOS DE LA COMPARACIÓN:")
    print("-" * 60)
    
    conceptos = [
        ('CRÉDITOS ASISTENCIALES', 'creditos_asistenciales'),
        ('FONDO VOLUNTARIO', 'fondo_voluntario'),
        ('PERSONAL', 'personal'),
        ('ADHERENTE', 'adherente'),
        ('PATRONAL', 'patronal'),
        ('TOTAL GENERAL', 'total')
    ]
    
    diferencias_encontradas = False
    total_diferencia = 0.0
    
    for nombre_concepto, clave in conceptos:
        csv_val = sumatorias_csv.get(clave, 0.0)
        acumulado_val = sumatorias_acumuladas.get(clave, 0.0)
        diferencia = abs(csv_val - acumulado_val)
        
        # Formatear valores para mostrar
        csv_formatted = f"${csv_val:,.2f}"
        acumulado_formatted = f"${acumulado_val:,.2f}"
        diferencia_formatted = f"${diferencia:,.2f}"
        
        # Determinar si hay diferencia significativa
        es_significativa = diferencia > 1.0  # Más de $1 de diferencia
        
        print(f"\n{nombre_concepto}:")
        print(f"  📄 CSV:               {csv_formatted:>20}")
        print(f"  📊 Acumulado tipos:   {acumulado_formatted:>20}")
        print(f"  ⚖  Diferencia:        {diferencia_formatted:>20}")
        
        if es_significativa:
            print(f"  ❌ DIFERENCIA SIGNIFICATIVA!")
            diferencias_encontradas = True
            total_diferencia += diferencia
            
            # Calcular porcentaje de diferencia
            if csv_val > 0:
                porcentaje = (diferencia / csv_val) * 100
                print(f"  📉 Porcentaje:        {porcentaje:.2f}%")
        else:
            print(f"  ✅ OK (diferencia < $1)")
    
    # Mostrar desglose por tipo
    print(f"\n{'='*60}")
    print("📋 DESGLOSE DETALLADO POR TIPO DE ENTIDAD:")
    print("-" * 60)
    
    for tipo_entidad, sumatorias in sumatorias_por_tipo.items():
        if tipo_entidad != 'Otro' and sumatorias['total'] > 0:
            print(f"\n🏢 {tipo_entidad.upper()}:")
            print(f"  │")
            print(f"  ├─ Créditos Asistenciales: ${sumatorias['creditos_asistenciales']:,.2f}")
            print(f"  ├─ Fondo Voluntario:       ${sumatorias['fondo_voluntario']:,.2f}")
            print(f"  ├─ Personal:               ${sumatorias['personal']:,.2f}")
            print(f"  ├─ Adherente:              ${sumatorias['adherente']:,.2f}")
            print(f"  ├─ Patronal:               ${sumatorias['patronal']:,.2f}")
            print(f"  └─ TOTAL:                 ${sumatorias['total']:,.2f}")
    
    # Mostrar resumen de tipos 'Otro'
    otros_total = sumatorias_por_tipo.get('Otro', {}).get('total', 0.0)
    if otros_total > 0:
        print(f"\n📌 NOTA: Hay ${otros_total:,.2f} en archivos clasificados como 'Otro'")
        print("  Estos archivos no fueron incluidos en el acumulado por tipo")
    
    # Mostrar estadísticas finales
    print(f"\n{'='*60}")
    print("📊 ESTADÍSTICAS FINALES:")
    print("-" * 60)
    
    print(f"Tipos de entidad con datos: {tipos_con_datos}")
    print(f"Total CSV: ${sumatorias_csv['total']:,.2f}")
    print(f"Total acumulado por tipos: ${sumatorias_acumuladas['total']:,.2f}")
    print(f"Diferencia total: ${total_diferencia:,.2f}")
    
    if not diferencias_encontradas:
        print(f"\n✅ TODAS LAS SUMATORIAS SON CONSISTENTES PARA PERÍODO {periodo}")
    else:
        print(f"\n⚠️  SE ENCONTRARON DIFERENCIAS EN PERÍODO {periodo}")
        print(f"   Total diferencia: ${total_diferencia:,.2f}")
        
        # Recomendaciones
        print(f"\n💡 RECOMENDACIONES:")
        print(f"   1. Revisar archivos clasificados como 'Otro'")
        print(f"   2. Verificar si algún archivo no se procesó completamente")
        print(f"   3. Revisar la extracción de datos de archivos problemáticos")
    
    print(f"{'='*70}")
    
    return not diferencias_encontradas, total_diferencia

def calcular_sumatorias_datos(datos_excel):
    """
    Calcula sumatorias directamente desde los datos extraídos.
    
    CORRECCIONES:
    - Personal: I (9-aporte personal) + Q (17-reajs aporte pers)
    - Adherente: J (10-adherente sec) + L (12-hijo menor de 35) + M (13-menor a cargo) + 
                R (18-reaj adherente sec) + T (20-reajuste hijo menor) + U (21-reajuste menor a cargo)
    - Fondo Voluntario: K (11-fondo v) + S (19-reajuste fv)
    - Créditos Asistenciales: N (14-cred asist) + V (22-reajuste cred asistencial)
    - Patronal: W (23-aporte patronal) + X (24-reajuste aporte patronal)
    """
    sumatorias = {
        'creditos_asistenciales': 0.0,
        'fondo_voluntario': 0.0,
        'personal': 0.0,
        'adherente': 0.0,
        'patronal': 0.0,
        'total': 0.0
    }
    
    filas_procesadas = 0
    filas_con_errores = 0
    
    for fila_idx, fila in enumerate(datos_excel, 1):
        try:
            # Verificar que la fila tenga al menos 24 columnas (ahora 25 con el código)
            if len(fila) < 24:
                filas_con_errores += 1
                if fila_idx <= 10:  # Solo mostrar primeros errores
                    print(f"   ⚠ Fila {fila_idx} tiene solo {len(fila)} columnas (necesita 24 para cálculos)")
                continue
            
            def safe_float(val):
                if not val or val == '' or str(val).strip() == '':
                    return 0.0
                try:
                    return float(str(val).strip())
                except ValueError:
                    return 0.0
            
            # OBTENER VALORES - CÁLCULOS CORREGIDOS
            # Columna I: 9-aporte personal
            aporte_personal = safe_float(fila[8])      
            
            # Columna J: 10-adherente sec
            adherente_sec = safe_float(fila[9])        
            
            # Columna K: 11-fondo v
            fondo_v = safe_float(fila[10])             
            
            # Columna L: 12-hijo menor de 35
            hijo_menor_35 = safe_float(fila[11])
            
            # Columna M: 13-menor a cargo
            menor_cargo = safe_float(fila[12])
            
            # Columna N: 14-cred asist
            cred_asist = safe_float(fila[13])          
            
            # Columna Q: 17-reajs aporte pers
            reaj_aporte_pers = safe_float(fila[16])    
            
            # Columna R: 18-reaj adherente sec
            reaj_adherente_sec = safe_float(fila[17])  
            
            # Columna S: 19-reajuste fv
            reajuste_fv = safe_float(fila[18])         
            
            # Columna T: 20-reajuste hijo menor
            reajuste_hijo_menor = safe_float(fila[19])
            
            # Columna U: 21-reajuste menor a cargo
            reajuste_menor_cargo = safe_float(fila[20])
            
            # Columna V: 22-reajuste cred asistencial
            reaj_cred_asist = safe_float(fila[21])     
            
            # Columna W: 23-aporte patronal
            aporte_patronal = safe_float(fila[22])     
            
            # Columna X: 24-reajuste aporte patronal
            reaj_aporte_patronal = safe_float(fila[23]) 
            
            # DEBUG: Mostrar primeras filas
            if fila_idx <= 3:
                print(f"\n🔍 DEBUG CÁLCULOS - Fila {fila_idx}:")
                print(f"  Col I (aporte personal): {aporte_personal:.2f}")
                print(f"  Col Q (reajuste aporte pers): {reaj_aporte_pers:.2f}")
                print(f"  → Personal fila: {(aporte_personal + reaj_aporte_pers):.2f}")
                
                print(f"  Col J (adherente sec): {adherente_sec:.2f}")
                print(f"  Col L (hijo menor 35): {hijo_menor_35:.2f}")
                print(f"  Col M (menor cargo): {menor_cargo:.2f}")
                print(f"  Col R (reaj adherente): {reaj_adherente_sec:.2f}")
                print(f"  Col T (reaj hijo menor): {reajuste_hijo_menor:.2f}")
                print(f"  Col U (reaj menor cargo): {reajuste_menor_cargo:.2f}")
                print(f"  → Adherente fila: {(adherente_sec + hijo_menor_35 + menor_cargo + reaj_adherente_sec + reajuste_hijo_menor + reajuste_menor_cargo):.2f}")
                
                # Mostrar código si está disponible
                if len(fila) > 24:
                    print(f"  Col 25 (Código): {fila[24]}")
            
            # CALCULAR SUMS POR CONCEPTO - FÓRMULAS CORREGIDAS
            # Personal = columna I + columna Q
            sum_personal_fila = aporte_personal + reaj_aporte_pers
            sumatorias['personal'] += sum_personal_fila
            
            # Adherente = columna J + columna L + columna M + columna R + columna T + columna U
            sum_adherente_fila = (
                adherente_sec + 
                hijo_menor_35 + 
                menor_cargo + 
                reaj_adherente_sec + 
                reajuste_hijo_menor + 
                reajuste_menor_cargo
            )
            sumatorias['adherente'] += sum_adherente_fila
            
            # Fondo Voluntario = columna K + columna S
            sum_fv_fila = fondo_v + reajuste_fv
            sumatorias['fondo_voluntario'] += sum_fv_fila
            
            # Créditos Asistenciales = columna N + columna V
            sum_creditos_fila = cred_asist + reaj_cred_asist
            sumatorias['creditos_asistenciales'] += sum_creditos_fila
            
            # Patronal = columna W + columna X
            sum_patronal_fila = aporte_patronal + reaj_aporte_patronal
            sumatorias['patronal'] += sum_patronal_fila
            
            filas_procesadas += 1
            
        except (ValueError, IndexError, TypeError) as e:
            filas_con_errores += 1
            if fila_idx <= 10:  # Solo mostrar primeros errores
                print(f"   ⚠ Error en fila {fila_idx}: {e}")
                print(f"     Fila: {fila}")
    
    # Calcular total general
    sumatorias['total'] = (
        sumatorias['personal'] + 
        sumatorias['adherente'] + 
        sumatorias['fondo_voluntario'] + 
        sumatorias['creditos_asistenciales'] + 
        sumatorias['patronal']
    )
    
    # Mostrar estadísticas de procesamiento
    if filas_con_errores > 0:
        print(f"\n📊 Estadísticas de cálculo de sumatorias:")
        print(f"  Filas procesadas correctamente: {filas_procesadas}")
        print(f"  Filas con errores: {filas_con_errores}")
        print(f"  Total filas: {len(datos_excel)}")
    
    return sumatorias

def extraer_y_preparar_datos_mes(drive, archivos_excel):
    """Extrae datos de archivos Excel y los prepara para el CSV"""
    mes_procesando = MES_ACTUAL

    datos_mes = []
    errores = []
    archivos_procesados = 0
    filas_totales = 0
    
    for archivo in archivos_excel:
        print(f"\n📄 Procesando: {archivo['name']}")
        
        try:
            # Descargar archivo
            fh = descargar_archivo(drive, archivo)
            if not fh:
                errores.append(f"No se pudo descargar: {archivo['name']}")
                continue
            
            # Extraer datos del Excel (solo los 24 campos A-X)
            datos_excel = extraer_datos_excel(fh, archivo['name'], mes_procesando)
            
            if datos_excel:
                # Agregar directamente los datos del Excel a datos_mes
                # Cada fila_excel ya contiene los 24 campos + código
                datos_mes.extend(datos_excel)
                
                filas_agregadas = len(datos_excel)
                filas_totales += filas_agregadas
                archivos_procesados += 1
                print(f"   ✅ {filas_agregadas} filas extraídas (24 campos + código = 25 columnas)")
            else:
                print(f"   ⚠ Sin datos en la hoja del mes {mes_procesando}")
                errores.append(f"Sin datos en hoja {mes_procesando}: {archivo['name']}")
                
        except Exception as e:
            error_msg = f"Error procesando {archivo['name']}: {str(e)}"
            print(f"   ❌ {error_msg}")
            errores.append(error_msg)
    
    return datos_mes, archivos_procesados, filas_totales, errores

def combinar_con_existente(csv_existente, datos_nuevos):
    """
    Combina CSV existente con nuevos datos.
    AHORA INCLUYE COLUMNA 25 PARA CÓDIGO.
    """
    if not csv_existente:
        # Si no hay CSV existente, crear uno nuevo con encabezados de 25 campos
        encabezados = [
            "1-cuil",
            "2-dni",
            "3-tipo doc",
            "4-nombre y apellido",
            "5-cod liq",
            "6-sit revista",
            "7-estado del afil",
            "8-reparticion",
            "9-aporte personal",
            "10-adherente sec",
            "11-fondo v",
            "12-hijo menor de 35",
            "13-menor a cargo",
            "14-cred asist",
            "15-sueldo sin desc",
            "16-sueldo con desc",
            "17-reajs aporte pers",
            "18-reaj adherente sec",
            "19-reajuste fv",
            "20-reajuste hijo menor",
            "21-reajuste menor a cargo",
            "22-reajuste cred asistencial",
            "23-aporte patronal",
            "24-reajuste aporte patronal",
            "25-codigo"  # NUEVA COLUMNA
        ]
        return encabezados, datos_nuevos
    
    # Leer CSV existente
    reader = csv.reader(io.StringIO(csv_existente))
    filas_existentes = list(reader)
    
    if not filas_existentes:
        return [], datos_nuevos
    
    # Separar encabezados y datos
    encabezados = filas_existentes[0]
    
    # Si el CSV existente no tiene columna 25, agregarla
    if len(encabezados) < 25:
        encabezados.append("25-codigo")
    
    datos_existentes = filas_existentes[1:]
    
    # Asegurar que los datos existentes tengan 25 columnas
    datos_existentes_normalizados = []
    for fila in datos_existentes:
        if len(fila) < 25:
            # Completar con celdas vacías hasta llegar a 25
            fila_completa = fila + [""] * (25 - len(fila))
            datos_existentes_normalizados.append(fila_completa)
        else:
            datos_existentes_normalizados.append(fila)
    
    # Combinar todos los datos existentes con nuevos datos
    datos_combinados = datos_existentes_normalizados + datos_nuevos
    
    return encabezados, datos_combinados

def determina_mes_a_procesar(mes_actual):
    
    if mes_actual == "12":
        return ["12", "2º sac"]
    
    if mes_actual == "06":
        return ["06", "1º sac"]
    
    else:
        return [mes_actual]

def extraer_y_preparar_datos_mes_periodo(drive, archivos_excel, periodo):
    """
    Versión modificada que verifica consistencia
    AHORA INCLUYE CÓDIGO EN LOS DATOS EXTRAÍDOS.
    """
    datos_mes = []
    errores = []
    archivos_procesados = 0
    filas_totales = 0
    
    # Diccionario para sumatorias por tipo de entidad - AGREGAR 'Otros'
    sumatorias_por_tipo = {
        'Municipios': {
            'creditos_asistenciales': 0.0,
            'fondo_voluntario': 0.0,
            'personal': 0.0,
            'adherente': 0.0,
            'patronal': 0.0,
            'total': 0.0
        },
        'Comunas': {
            'creditos_asistenciales': 0.0,
            'fondo_voluntario': 0.0,
            'personal': 0.0,
            'adherente': 0.0,
            'patronal': 0.0,
            'total': 0.0
        },
        'Entes Descentralizados': {
            'creditos_asistenciales': 0.0,
            'fondo_voluntario': 0.0,
            'personal': 0.0,
            'adherente': 0.0,
            'patronal': 0.0,
            'total': 0.0
        },
        'Cajas Municipales': {
            'creditos_asistenciales': 0.0,
            'fondo_voluntario': 0.0,
            'personal': 0.0,
            'adherente': 0.0,
            'patronal': 0.0,
            'total': 0.0
        },
        'Escuela': {
            'creditos_asistenciales': 0.0,
            'fondo_voluntario': 0.0,
            'personal': 0.0,
            'adherente': 0.0,
            'patronal': 0.0,
            'total': 0.0
        },
        'Otros': {  # AGREGAR ESTA CATEGORÍA
            'creditos_asistenciales': 0.0,
            'fondo_voluntario': 0.0,
            'personal': 0.0,
            'adherente': 0.0,
            'patronal': 0.0,
            'total': 0.0
        }
    }
    
    # Variables para totales directos del período
    sumatorias_directas_periodo = {
        'creditos_asistenciales': 0.0,
        'fondo_voluntario': 0.0,
        'personal': 0.0,
        'adherente': 0.0,
        'patronal': 0.0,
        'total': 0.0
    }
    
    # Contadores para estadísticas
    archivos_por_tipo = {}
    
    for archivo in archivos_excel:
        print(f"\n📄 Procesando: {archivo['name']} (período: {periodo})")
        
        try:
            # Determinar tipo de entidad (USANDO LA FUNCIÓN MEJORADA)
            tipo_entidad = determinar_tipo_reparticion(archivo['name'])
            
            # Actualizar contador de archivos por tipo
            archivos_por_tipo[tipo_entidad] = archivos_por_tipo.get(tipo_entidad, 0) + 1
            
            # DEBUG: Mostrar tipo para verificar
            if 'municipio' in archivo['name'].lower() or 'municipal' in archivo['name'].lower():
                print(f"   🏢 ARCHIVO MUNICIPIO DETECTADO: {archivo['name']} -> {tipo_entidad}")
            
            # Descargar archivo
            fh = descargar_archivo(drive, archivo)
            if not fh:
                errores.append(f"No se pudo descargar: {archivo['name']}")
                continue
            
            # Extraer datos del Excel para el período específico
            datos_excel = extraer_datos_excel(fh, archivo['name'], periodo)
            
            if datos_excel:
                # Calcular sumatorias para este archivo
                sumatorias_archivo = calcular_sumatorias_datos(datos_excel)
                
                # DEBUG DETALLADO PARA ARCHIVOS IMPORTANTES
                if tipo_entidad == 'Municipios' or 'municipio' in archivo['name'].lower():
                    print(f"\n   💰 DEBUG DETALLADO - Archivo: {archivo['name']}")
                    print(f"      Tipo: {tipo_entidad}")
                    print(f"      Filas extraídas: {len(datos_excel)}")
                    print(f"      Personal calculado: ${sumatorias_archivo['personal']:,.2f}")
                    print(f"      Adherente calculado: ${sumatorias_archivo['adherente']:,.2f}")
                    print(f"      Fondo V calculado: ${sumatorias_archivo['fondo_voluntario']:,.2f}")
                    print(f"      Créditos calculado: ${sumatorias_archivo['creditos_asistenciales']:,.2f}")
                    print(f"      Patronal calculado: ${sumatorias_archivo['patronal']:,.2f}")
                    print(f"      TOTAL archivo: ${sumatorias_archivo['total']:,.2f}")
                
                # Acumular sumatorias por tipo de entidad
                for concepto in sumatorias_archivo:
                    if concepto in sumatorias_por_tipo[tipo_entidad]:
                        sumatorias_por_tipo[tipo_entidad][concepto] += sumatorias_archivo[concepto]
                        sumatorias_directas_periodo[concepto] += sumatorias_archivo[concepto]
                
                datos_mes.extend(datos_excel)
                filas_agregadas = len(datos_excel)
                filas_totales += filas_agregadas
                archivos_procesados += 1
                print(f"   ✅ {filas_agregadas} filas extraídas del período {periodo} (con código en columna 25)")
                
                # Mostrar avance cada 5 archivos
                if archivos_procesados % 5 == 0:
                    print(f"   📈 Progreso: {archivos_procesados}/{len(archivos_excel)} archivos procesados")
                
            else:
                print(f"   ⚠ Sin datos en la hoja del período {periodo}")
                errores.append(f"Sin datos en período {periodo}: {archivo['name']}")
                
        except Exception as e:
            error_msg = f"Error procesando {archivo['name']} (período {periodo}): {str(e)}"
            print(f"   ❌ {error_msg}")
            print(f"   Traceback: {traceback.format_exc()}")
            errores.append(error_msg)
    
    # Mostrar estadísticas de archivos por tipo
    print(f"\n📊 ESTADÍSTICAS DE ARCHIVOS POR TIPO - PERÍODO {periodo}:")
    print("-" * 50)
    
    # Filtrar items no nulos y ordenar
    items_validos = [(tipo, cantidad) for tipo, cantidad in archivos_por_tipo.items() if tipo is not None]
    items_ordenados = sorted(items_validos, key=lambda x: x[0] if x[0] else "")  # Ordenar por tipo
    
    for tipo, cantidad in items_ordenados:
        print(f"  {tipo}: {cantidad} archivo(s)")
    
    # Si hay archivos sin tipo (None), mostrarlos por separado
    archivos_sin_tipo = sum(cantidad for tipo, cantidad in archivos_por_tipo.items() if tipo is None)
    if archivos_sin_tipo > 0:
        print(f"  Sin tipo asignado: {archivos_sin_tipo} archivo(s)")
    
    # Mostrar sumatorias finales por tipo de entidad para este período
    print(f"\n📊 RESUMEN SUMATORIAS POR TIPO DE ENTIDAD - PERÍODO {periodo}:")
    print("-" * 60)
    
    total_por_tipos = 0.0
    for tipo_entidad, sumatorias in sorted(sumatorias_por_tipo.items()):
        if sumatorias['total'] > 0:
            total_por_tipos += sumatorias['total']
            print(f"\n  🏢 {tipo_entidad.upper()}:")
            print(f"    ├─ Créditos Asistenciales: ${sumatorias['creditos_asistenciales']:,.2f}")
            print(f"    ├─ Fondo Voluntario:       ${sumatorias['fondo_voluntario']:,.2f}")
            print(f"    ├─ Personal:               ${sumatorias['personal']:,.2f}")
            print(f"    ├─ Adherente:              ${sumatorias['adherente']:,.2f}")
            print(f"    ├─ Patronal:               ${sumatorias['patronal']:,.2f}")
            print(f"    └─ TOTAL {tipo_entidad}:     ${sumatorias['total']:,.2f}")
    
    print(f"\n  📈 TOTAL POR TIPOS: ${total_por_tipos:,.2f}")
    
    # Mostrar totales directos del período
    print(f"\n📈 TOTALES DIRECTOS CALCULADOS - PERÍODO {periodo}:")
    print("-" * 50)
    for concepto, valor in sumatorias_directas_periodo.items():
        print(f"  {concepto.replace('_', ' ').title()}: ${valor:,.2f}")
    
    return datos_mes, archivos_procesados, filas_totales, errores, sumatorias_por_tipo, sumatorias_directas_periodo

def ejecutar_principal():
    """Función principal del unificador mensual"""
    import traceback
    
    inicio = time.time()
    mes_actual = MES_ACTUAL
    anio_actual = obtener_anio(mes_actual)
    
    periodos = determina_mes_a_procesar(mes_actual)
   
    ahora = registrar_inicio(f"UNIFICADOR MENSUAL - PROCESANDO {len(periodos)} PERÍODO(S)")
    
    # 1. Inicializar Drive
    drive = inicializar_drive()
    if not drive:
        print("❌ No se pudo inicializar Drive")
        return
    
    # 2. Obtener archivos Excel
    print("📁 Buscando archivos Excel en Drive...")
    archivos = obtener_archivos(drive)
    
    # Filtrar solo Excel
    archivos_excel = []
    for a in archivos:
        nombre = a["name"].lower()
        mime = a["mimeType"]
        
        es_excel = (
            nombre.endswith(".xlsx") or
            nombre.endswith(".xlsm") or
            nombre.endswith(".xls") or
            mime == "application/vnd.google-apps.spreadsheet"
        )
        
        if es_excel:
            archivos_excel.append(a)
    
    print(f"✅ Archivos Excel válidos: {len(archivos_excel)}")
    
    # DEBUG: Mostrar nombres de archivos para verificar clasificación
    print(f"\n📋 LISTA DE ARCHIVOS EXCEL ENCONTRADOS ({len(archivos_excel)}):")
    
    
    archivos_csv_generados = []
    total_filas_todos_periodos = 0
    todos_errores = []
    cantidades_por_periodo = {}
    
    # Diccionario para acumular sumatorias por tipo de entidad para todos los períodos
    sumatorias_totales_por_tipo = {
        'Municipios': {
            'creditos_asistenciales': 0.0,
            'fondo_voluntario': 0.0,
            'personal': 0.0,
            'adherente': 0.0,
            'patronal': 0.0,
            'total': 0.0
        },
        'Comunas': {
            'creditos_asistenciales': 0.0,
            'fondo_voluntario': 0.0,
            'personal': 0.0,
            'adherente': 0.0,
            'patronal': 0.0,
            'total': 0.0
        },
        'Entes Descentralizados': {
            'creditos_asistenciales': 0.0,
            'fondo_voluntario': 0.0,
            'personal': 0.0,
            'adherente': 0.0,
            'patronal': 0.0,
            'total': 0.0
        },
        'Cajas Municipales': {
            'creditos_asistenciales': 0.0,
            'fondo_voluntario': 0.0,
            'personal': 0.0,
            'adherente': 0.0,
            'patronal': 0.0,
            'total': 0.0
        },
        'Escuela': {
            'creditos_asistenciales': 0.0,
            'fondo_voluntario': 0.0,
            'personal': 0.0,
            'adherente': 0.0,
            'patronal': 0.0,
            'total': 0.0
        }
    }
    
    sumatorias_por_periodo_y_tipo = {}
    # Nuevo diccionario para guardar sumatorias directas por período
    sumatorias_directas_por_periodo = {}
    
    # Estadísticas de consistencia
    consistencias_por_periodo = {}
    diferencias_totales = {}

    # 3. Procesar cada período por separado
    for periodo in periodos:
        print(f"\n{'='*70}")
        print(f"🔄 PROCESANDO PERÍODO: {periodo}")
        print(f"{'='*70}")
        
        nombre_periodo = nombre_mes(periodo)

        # 4. Extraer datos de este período específico con sumatorias directas
        datos_periodo, archivos_procesados, filas_periodo, errores, sumatorias_por_tipo, sumatorias_directas = extraer_y_preparar_datos_mes_periodo(
            drive, archivos_excel, periodo
        )
        
        # Guardar las sumatorias
        sumatorias_por_periodo_y_tipo[periodo] = sumatorias_por_tipo
        sumatorias_directas_por_periodo[periodo] = sumatorias_directas
        
        # Acumular sumatorias totales
        for tipo_entidad in sumatorias_por_tipo:
            if tipo_entidad in sumatorias_totales_por_tipo:
                for concepto in sumatorias_por_tipo[tipo_entidad]:
                    if concepto in sumatorias_totales_por_tipo[tipo_entidad]:
                        sumatorias_totales_por_tipo[tipo_entidad][concepto] += sumatorias_por_tipo[tipo_entidad][concepto]
        
        # Guardar la cantidad para este período
        cantidades_por_periodo[periodo] = len(datos_periodo)

        if not datos_periodo:
            print(f"⚠️ No se extrajeron datos para el período {periodo}")
            todos_errores.extend(errores)
            consistencias_por_periodo[periodo] = False
            diferencias_totales[periodo] = 0.0
            continue
        
        # 5. Crear encabezados para este período (AHORA CON 25 COLUMNAS)
        encabezados = [
                "1-cuil", "2-dni", "3-tipo doc", "4-nombre y apellido", "5-cod liq",
                "6-sit revista", "7-estado del afil", "8-reparticion", "9-aporte personal",
                "10-adherente sec", "11-fondo v", "12-hijo menor de 35", "13-menor a cargo",
                "14-cred asist", "15-sueldo sin desc", "16-sueldo con desc", "17-reajs aporte pers",
                "18-reaj adherente sec", "19-reajuste fv", "20-reajuste hijo menor",
                "21-reajuste menor a cargo", "22-reajuste cred asistencial", "23-aporte patronal",
                "24-reajuste aporte patronal", "25-codigo"
            ]
        
        # Agregar encabezados a los datos
        datos_finales = [encabezados] + datos_periodo
        
        # 6. Generar nombre del CSV para este período
        nombre_csv = f"Unificado_{nombre_periodo}{anio_actual}.csv"
        
        # 7. Guardar CSV localmente
        print(f"\n💾 Guardando CSV para período {periodo}...")
        ruta_csv_local = guardar_csv_localmente(datos_finales, nombre_csv)
        
        if ruta_csv_local:
            archivos_csv_generados.append(ruta_csv_local)
            total_filas_todos_periodos += len(datos_periodo)
            
            print(f"✅ CSV guardado: {nombre_csv}")
            print(f"📊 Filas en este período: {len(datos_periodo)}")
            print(f"📋 Columnas totales: 25 (incluye código en columna 25)")
            
            # 8. VERIFICAR CONSISTENCIA ENTRE DATOS DIRECTOS Y CSV
            print(f"\n🔍 EJECUTANDO VERIFICACIÓN DETALLADA DE CONSISTENCIA...")
            
            # Ejecutar verificación detallada
            es_consistente, diferencia_total = verificar_consistencia_sumatorias_detallada(
                periodo, 
                sumatorias_por_tipo,
                ruta_csv_local
            )
            
            consistencias_por_periodo[periodo] = es_consistente
            diferencias_totales[periodo] = diferencia_total
            
            # Verificar el tamaño del archivo
            if os.path.exists(ruta_csv_local):
                file_size = os.path.getsize(ruta_csv_local) / (1024 * 1024)  # MB
                print(f"📦 Tamaño del archivo: {file_size:.2f} MB")
        else:
            error_msg = f"No se pudo guardar CSV para período {periodo}"
            print(f"❌ {error_msg}")
            todos_errores.append(error_msg)
            consistencias_por_periodo[periodo] = False
            diferencias_totales[periodo] = 0.0
    
    # 9. Mostrar resumen final de consistencia
    print(f"\n{'='*70}")
    print("📊 RESUMEN FINAL DE CONSISTENCIA")
    print(f"{'='*70}")
    
    periodos_consistentes = 0
    periodos_inconsistentes = 0
    total_diferencia_acumulada = 0.0
    
    for periodo in periodos:
        nombre_periodo = nombre_mes(periodo)
        es_consistente = consistencias_por_periodo.get(periodo, False)
        diferencia = diferencias_totales.get(periodo, 0.0)
        
        if es_consistente:
            print(f"✅ PERÍODO {periodo} ({nombre_periodo}): CONSISTENTE (diferencia: ${diferencia:,.2f})")
            periodos_consistentes += 1
        else:
            print(f"❌ PERÍODO {periodo} ({nombre_periodo}): INCONSISTENTE (diferencia: ${diferencia:,.2f})")
            periodos_inconsistentes += 1
        
        total_diferencia_acumulada += diferencia
    
    print(f"\n📈 ESTADÍSTICAS DE CONSISTENCIA:")
    print(f"  Períodos consistentes: {periodos_consistentes}/{len(periodos)}")
    print(f"  Períodos inconsistentes: {periodos_inconsistentes}/{len(periodos)}")
    print(f"  Diferencia total acumulada: ${total_diferencia_acumulada:,.2f}")
    
    # 10. Mostrar resumen final de sumatorias
    print(f"\n{'='*70}")
    print("📊 SUMATORIAS TOTALES POR TIPO DE ENTIDAD (TODOS LOS PERÍODOS):")
    print(f"{'='*70}")
    
    total_general = 0.0
    for tipo_entidad, sumatorias in sumatorias_totales_por_tipo.items():
        if sumatorias['total'] > 0:
            total_general += sumatorias['total']
            print(f"\n🏢 {tipo_entidad.upper()}:")
            print(f"  ├─ Créditos Asistenciales: ${sumatorias['creditos_asistenciales']:,.2f}")
            print(f"  ├─ Fondo Voluntario:       ${sumatorias['fondo_voluntario']:,.2f}")
            print(f"  ├─ Personal:               ${sumatorias['personal']:,.2f}")
            print(f"  ├─ Adherente:              ${sumatorias['adherente']:,.2f}")
            print(f"  ├─ Patronal:               ${sumatorias['patronal']:,.2f}")
            print(f"  └─ TOTAL {tipo_entidad}:     ${sumatorias['total']:,.2f}")
    
    print(f"\n{'='*50}")
    print(f"💰 TOTAL GENERAL: ${total_general:,.2f}")
    print(f"{'='*50}")
    
    # 11. Mostrar resumen del proceso
    registrar_resumen(
        inicio,
        archivos_procesados=len(archivos_excel),
        archivos_encontrados=len(archivos_excel),
        filas_procesadas=total_filas_todos_periodos,
        errores=todos_errores
    )
    
    # 12. Enviar email con todos los archivos adjuntos
    print("\n📧 Preparando email de resumen...")
    
    nombre_del_mes = nombre_mes(MES_ACTUAL)
    periodo_legible_upper = f"{nombre_del_mes.upper()}/{anio_actual}"
    periodo_s = "PERIODO"

    # Preparar asunto del email
    if len(periodos) == 2:
        # Para el caso Junio+1°sac y Diciembre+2°sac
        periodo_s = "PERIODOS"
        if MES_ACTUAL == "06":
            periodo_legible_upper += f" y 1° SAC/{anio_actual}"
        
        elif MES_ACTUAL == "12":
            periodo_legible_upper += f" y 2° SAC/{anio_actual}"
    
    asunto = f"🟢🔵 OSER - UNIFICADO MENSUAL AUTOMÁTICO | {periodo_s}: {periodo_legible_upper}"
    
    # Generar HTML con resumen de todos los períodos INCLUYENDO DESGLOSE POR TIPO
    html = generar_html_resumen_unificador(
        periodos,
        ahora.strftime("%d-%m-%Y %H:%M:%S"),
        cantidades_por_periodo,
        anio_actual,
        sumatorias_por_periodo_y_tipo
    )
    
    # Enviar email con todos los archivos CSV adjuntos
    print(f"📎 Adjuntando {len(archivos_csv_generados)} archivo(s) CSV al email...")
    
    # Filtrar archivos que existen y tienen tamaño razonable (<25MB)
    adjuntos_validos = []
    for ruta in archivos_csv_generados:
        if os.path.exists(ruta):
            file_size = os.path.getsize(ruta) / (1024 * 1024)
            if file_size < 25:
                adjuntos_validos.append(ruta)
                print(f"  ✅ {os.path.basename(ruta)} ({file_size:.2f} MB)")
            else:
                print(f"  ⚠️  {os.path.basename(ruta)} demasiado grande ({file_size:.2f} MB) - no se adjunta")
        else:
            print(f"  ❌ {os.path.basename(ruta)} no encontrado")
    
    enviar_email_html_con_adjuntos(asunto, html, adjuntos_validos, "SMTP_TO_UNIFICADOR")
    
    print("\n" + "=" * 70)
    print("✅ PROCESO COMPLETADO!")
    print("=" * 70)
    print(f"📁 Archivos generados: {len(archivos_csv_generados)}")
    for i, ruta in enumerate(archivos_csv_generados, 1):
        nombre = os.path.basename(ruta)
        if os.path.exists(ruta):
            filas = sum(1 for _ in open(ruta, 'r', encoding='utf-8')) - 1  # Excluir encabezado
            print(f"  {i}. {nombre} - {filas} filas")
    print(f"📊 Total de filas combinadas: {total_filas_todos_periodos}")
    print(f"📧 Email enviado con {len(adjuntos_validos)} adjunto(s)")
    print("=" * 70)


if __name__ == "__main__":
    ejecutar_principal()