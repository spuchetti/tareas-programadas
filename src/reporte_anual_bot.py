"""
Bot de Reporte Anual

Consolida los datos del unificador mensual para los 12 meses
del año en un unico Excel, con una hoja por periodo.

"""
"""
======= EJECUCION MANUAL =========
Usar workflow_dispatch en GitHub Actions con el input:
  - anio: ej: 2025

"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import sys
import os
import re
import traceback
import time
from datetime import datetime

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from utils.common_utils import (
    registrar_inicio, registrar_resumen,
    nombre_mes, crear_directorio_salida
)
from utils.drive_utils import inicializar_drive, obtener_archivos, descargar_archivo
from utils.excel_utils import normalizar_texto
from utils.gmail_utils import enviar_email_html_con_adjuntos, generar_html_resumen_anual


# ---------------------------------------------------------------------------
# Configuracion
# ---------------------------------------------------------------------------

_anio_override = os.getenv("ANIO_OVERRIDE", "").strip()
ANIO_ACTUAL = int(_anio_override)

MESES = ["01", "02", "03", "04", "05", "06", "1° sac", "07", "08", "09", "10", "11","2° sac", "12"]

ENCABEZADOS = [
    "1-cuil", "2-dni", "3-tipo doc", "4-nombre y apellido", "5-cod liq",
    "6-sit revista", "7-estado del afil", "8-reparticion", "9-aporte personal",
    "10-adherente sec", "11-fondo v", "12-hijo menor de 35", "13-menor a cargo",
    "14-cred asist", "15-sueldo sin desc", "16-sueldo con desc", "17-reajs aporte pers",
    "18-reaj adherente sec", "19-reajuste fv", "20-reajuste hijo menor",
    "21-reajuste menor a cargo", "22-reajuste cred asistencial", "23-aporte patronal",
    "24-reajuste aporte patronal", "25-codigo"
]

# Estilos de Excel
FILL_HEADER  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
FONT_HEADER  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
FONT_DATA    = Font(name="Arial", size=10)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center")


# ---------------------------------------------------------------------------
# Extraccion de datos
# ---------------------------------------------------------------------------

def extraer_codigo_desde_nombre(nombre_archivo):
    try:
        nombre_sin_ext = os.path.splitext(nombre_archivo)[0]
        partes = nombre_sin_ext.split('-', 1)
        codigo = partes[0].strip() if partes else ""
        return codigo if codigo else "SIN_CODIGO"
    except Exception:
        return "SIN_CODIGO"


def extraer_datos_excel(fh, nombre_archivo, hoja_mes):
    """
    Extrae columnas A-X (+ codigo como col 25) desde fila 4 (fila 5 para archivos Caja).
    Se detiene en '-' o celda vacia en columna A.
    Columnas D, F, H: elimina tildes. Cols I-X: formato numerico 2 decimales.
    """
    try:
        codigo_archivo = extraer_codigo_desde_nombre(nombre_archivo)

        fh.seek(0)
        wb = openpyxl.load_workbook(fh, data_only=True, read_only=True)

        if hoja_mes not in wb.sheetnames:
            wb.close()
            return []

        ws = wb[hoja_mes]
        es_caja   = "caja" in nombre_archivo.lower()
        fila_inicio = 5 if es_caja else 4
        datos = []

        for row in ws.iter_rows(min_row=fila_inicio, max_col=24, values_only=True):
            primera = row[0] if row else None
            if primera is None:
                primera_str = ""
            elif isinstance(primera, datetime):
                primera_str = primera.strftime("%Y-%m-%d")
            else:
                primera_str = str(primera).strip()

            if primera_str in ("-", ""):
                break
            if all(c is None or str(c).strip() == "" for c in row):
                break

            fila_limpia = []
            for col_idx, cell in enumerate(row, start=1):
                if cell is None:
                    fila_limpia.append("" if col_idx <= 8 else "0.00")
                elif isinstance(cell, datetime):
                    fila_limpia.append(cell.strftime("%Y-%m-%d"))
                elif col_idx <= 8:
                    # Columnas de texto
                    if isinstance(cell, (int, float)):
                        val = str(int(cell)) if isinstance(cell, float) and cell.is_integer() else str(cell)
                    else:
                        val = str(cell)
                    # D(4), F(6), H(8): elimina tildes
                    eliminar = col_idx in (4, 6, 8)
                    fila_limpia.append(normalizar_texto(val, eliminar_tildes_param=eliminar))
                else:
                    # Columnas numericas I-X
                    if isinstance(cell, (int, float)):
                        fila_limpia.append(f"{float(cell):.2f}")
                    else:
                        s = str(cell).strip()
                        if s in ("", "nan", "0"):
                            fila_limpia.append("0.00")
                        else:
                            try:
                                s = s.replace(',', '.')
                                if s.count('.') > 1:
                                    partes = s.split('.')
                                    s = ''.join(partes[:-1]) + '.' + partes[-1]
                                fila_limpia.append(f"{float(s):.2f}")
                            except ValueError:
                                fila_limpia.append("0.00")

            if any(c != "" and c != "0.00" for c in fila_limpia):
                datos.append(fila_limpia + [codigo_archivo])

        wb.close()
        print(f" ✔ {nombre_archivo} [{hoja_mes}]: {len(datos)} filas")
        return datos

    except Exception as e:
        print(f"✘ Error extrayendo {nombre_archivo}: {e}")
        traceback.print_exc()
        return []


# ---------------------------------------------------------------------------
# Generacion de Excel
# ---------------------------------------------------------------------------

ANCHOS_COL = [
    18, 14, 12, 30, 10,  # A-E
    20, 18, 35, 14, 14,  # F-J
    12, 16, 16, 14, 16,  # K-O
    16, 16, 18, 14, 18,  # P-T
    20, 22, 16, 22, 12   # U-Y (Y = codigo)
]


def aplicar_encabezados(ws):
    ws.append(ENCABEZADOS)
    for i, cell in enumerate(ws[1], start=1):
        cell.font      = FONT_HEADER
        cell.fill      = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        letra = cell.column_letter
        ws.column_dimensions[letra].width = ANCHOS_COL[i - 1] if i <= len(ANCHOS_COL) else 14
    ws.row_dimensions[1].height = 18


def agregar_filas_excel(ws, filas):
    fmt_num = '#,##0.00'
    for fila in filas:
        ws.append(fila)

    for row in ws.iter_rows(min_row=2):
        for i, cell in enumerate(row, start=1):
            cell.font = FONT_DATA
            if i <= 8 or i == 25:
                cell.alignment = ALIGN_LEFT
            else:
                cell.number_format = fmt_num
                cell.alignment     = ALIGN_CENTER
                # Convertir string numerico a float para que Excel lo trate como numero
                if cell.value and isinstance(cell.value, str):
                    try:
                        cell.value = float(cell.value)
                    except ValueError:
                        pass


def generar_excel_anual(datos_por_mes, anio):
    """
    Crea el workbook con una hoja por mes.
    datos_por_mes: dict { "01": [...filas...], ... }
    Retorna ruta del archivo o None.
    """
    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for mes in MESES:
            nombre_hoja = nombre_mes(mes)
            ws = wb.create_sheet(title=nombre_hoja)
            aplicar_encabezados(ws)
            filas = datos_por_mes.get(mes, [])
            agregar_filas_excel(ws, filas)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:Y1"
            print(f" Hoja '{nombre_hoja}': {len(filas)} registros")

        carpeta = crear_directorio_salida()
        nombre_archivo = f"Unificado_Anual_{anio}.xlsx"
        ruta = os.path.join(carpeta, nombre_archivo)
        wb.save(ruta)

        print(f"\n ✔ Excel generado: {nombre_archivo}  ({ruta})")
        return ruta

    except Exception as e:
        print(f"✘ Error generando Excel anual: {e}")
        traceback.print_exc()
        return None


# ---------------------------------------------------------------------------
# Principal
# ---------------------------------------------------------------------------

def ejecutar_principal():
    inicio = time.time()
    ahora  = registrar_inicio("ツ BOT REPORTE ANUAL UNIFICADO")

    drive = inicializar_drive()
    if not drive:
        return

    archivos = obtener_archivos(drive)
    if not archivos:
        print("✘ No se encontraron archivos.")
        return

    print(f" Año a procesar: {ANIO_ACTUAL}")
    print(f" Meses: {', '.join(MESES)}")
    print(f" Extrae columnas A-X + codigo, igual que el unificador mensual\n")

    datos_por_mes  = {mes: [] for mes in MESES}
    resumen_por_mes = {}

    for mes in MESES:
        nombre_legible = nombre_mes(mes)
        print(f"\n{'='*60}")
        print(f" Procesando mes: {nombre_legible} ({mes}/{ANIO_ACTUAL})")
        print(f"{'='*60}")

        filas_mes = []
        archivos_con_datos = 0

        for archivo in archivos:
            fh = descargar_archivo(drive, archivo)
            if not fh:
                continue
            filas = extraer_datos_excel(fh, archivo["name"], mes)
            if filas:
                filas_mes.extend(filas)
                archivos_con_datos += 1

        datos_por_mes[mes] = filas_mes
        resumen_por_mes[mes] = {
            "nombre":    nombre_legible,
            "registros": len(filas_mes),
            "archivos":  archivos_con_datos,
        }
        print(f"   → {len(filas_mes)} registros en {archivos_con_datos} reparticion(es)")

    # Genera Excel
    print(f"\n{'='*60}")
    print("Generando archivo Excel anual...")
    ruta_excel = generar_excel_anual(datos_por_mes, ANIO_ACTUAL)

    # Email
    if ruta_excel:
        html = generar_html_resumen_anual(
            ANIO_ACTUAL,
            len(archivos),
            resumen_por_mes,
            ahora.strftime("%d-%m-%Y %H:%M:%S")
        )
        asunto = f"🟢🔵 OSER - UNIFICADO ANUAL | AÑO: {ANIO_ACTUAL}"
        enviar_email_html_con_adjuntos(asunto, html, [ruta_excel], "SMTP_TO_UNIFICADOR")
    else:
        print("⚠ No se genero el Excel, email no enviado.")

    # Resumen consola
    print(f"\n{'='*70}")
    print(" ⠿ RESUMEN POR MES")
    print(f"{'='*70}")
    for mes in MESES:
        r = resumen_por_mes[mes]
        print(f"  {r['nombre']:12s}: {r['registros']:5d} registros  ({r['archivos']} archivos)")
    total = sum(v["registros"] for v in resumen_por_mes.values())
    print(f"{'─'*70}")
    print(f"  {'TOTAL':12s}: {total:5d} registros")
    print(f"{'='*70}")

    registrar_resumen(inicio, len(archivos), len(archivos))


if __name__ == "__main__":
    ejecutar_principal()